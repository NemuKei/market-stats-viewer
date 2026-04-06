from openpyxl import Workbook

from scripts.update_data import find_monthly_sheet_name, find_ts_table_xlsx_url


def test_find_ts_table_xlsx_url_prefers_pushi_hyo_link() -> None:
    html = """
    <html>
      <body>
        <a href="/content/001992573.xlsx">2026年2月分（第1次速報値）集計結果</a>
        <a href="/content/001992670.xlsx">推移表［Excel:1.3MB］</a>
      </body>
    </html>
    """

    url = find_ts_table_xlsx_url(
        html,
        "https://www.mlit.go.jp/kankocho/tokei_hakusyo/shukuhakutokei.html",
    )

    assert url == "https://www.mlit.go.jp/content/001992670.xlsx"


def test_find_monthly_sheet_name_ignores_old_sheets() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "1-1"
    wb.create_sheet("旧1-1")
    wb.create_sheet("旧1-2")
    wb.create_sheet("2-2")
    wb.create_sheet("4-1")

    assert find_monthly_sheet_name(wb, "1") == "1-1"
    assert find_monthly_sheet_name(wb, "1", legacy=True) == "旧1-2"
    assert find_monthly_sheet_name(wb, "2") == "2-2"
    assert find_monthly_sheet_name(wb, "4") == "4-1"
