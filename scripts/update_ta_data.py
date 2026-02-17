from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup

TA_SOURCE_INDEX_URL = (
    "https://www.mlit.go.jp/kankocho/tokei_hakusyo/ryokogyotoriatsukaigaku.html"
)

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data"
SQLITE_PATH = DATA_DIR / "market_stats.sqlite"
META_PATH = DATA_DIR / "meta_ta.json"
TABLE_NAME = "ta_company_amounts"


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = normalize_text(value).replace(",", "")
    if not s or s in {"-", "－", "…"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def save_meta(meta: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    META_PATH.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def fetch_html(url: str) -> str:
    res = requests.get(url, timeout=60)
    res.raise_for_status()
    if not res.encoding or res.encoding.lower() == "iso-8859-1":
        res.encoding = res.apparent_encoding
    return res.text


def download_file(url: str, dst: Path) -> None:
    with requests.get(url, stream=True, timeout=120) as res:
        res.raise_for_status()
        with dst.open("wb") as f:
            for chunk in res.iter_content(chunk_size=1024 * 64):
                if chunk:
                    f.write(chunk)


def parse_fiscal_year(text: str) -> str | None:
    m = re.search(r"\((20\d{2})年度\)", text)
    if m:
        return m.group(1)
    m = re.search(r"(20\d{2})年度", text)
    if m:
        return m.group(1)
    return None


def parse_period_from_link_text(link_text: str) -> str | None:
    text = normalize_text(link_text).replace(" ", "")
    if "総計" in text:
        return "total"
    m = re.search(r"(20\d{2})年.*?([0-9]{1,2})月", text)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        return f"{year:04d}-{month:02d}"
    return None


def extract_fiscal_page_links(html: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[dict[str, str]] = []
    seen: set[str] = set()

    for a in soup.find_all("a"):
        href = normalize_text(a.get("href"))
        text = normalize_text(a.get_text())
        if not href:
            continue

        abs_url = urljoin(base_url, href)
        if "ryokogyotoriatsukaigaku" not in abs_url:
            continue
        if not abs_url.lower().endswith(".html"):
            continue
        if "年度" not in text and "_r" not in abs_url:
            continue
        if abs_url in seen:
            continue

        seen.add(abs_url)
        out.append({"url": abs_url, "link_text": text})

    if not out:
        raise RuntimeError("No fiscal-year TA pages were found.")

    def sort_key(item: dict[str, str]) -> tuple[int, str]:
        y = parse_fiscal_year(item["link_text"]) or "0"
        return int(y), item["url"]

    return sorted(out, key=sort_key, reverse=True)


def extract_ta_excel_links(html: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[dict[str, str]] = []
    seen: set[str] = set()

    for a in soup.find_all("a"):
        href = normalize_text(a.get("href"))
        text = normalize_text(a.get_text())
        if not href:
            continue

        abs_url = urljoin(base_url, href)
        lower_url = abs_url.lower()
        if not (lower_url.endswith(".xlsx") or lower_url.endswith(".xls")):
            continue
        if "各社別内訳" not in text:
            continue
        if abs_url in seen:
            continue
        seen.add(abs_url)
        out.append({"url": abs_url, "link_text": text})

    return out


def classify_segment(header_text: str) -> str | None:
    compact = header_text.replace(" ", "").replace("　", "")
    if "外国人旅行" in compact:
        return "foreign"
    if "海外旅行" in compact:
        return "overseas"
    if "国内旅行" in compact:
        return "domestic"
    if "合計" in compact:
        return "total"
    return None


def find_amount_column(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_col: int,
    current_period_token: str,
) -> int | None:
    candidate_col: int | None = None
    for c in range(start_col, start_col + 5):
        year_label = normalize_text(ws.cell(row=5, column=c).value)
        unit_label = normalize_text(ws.cell(row=6, column=c).value)
        if "取扱額" not in unit_label:
            continue
        if candidate_col is None:
            candidate_col = c
        if year_label == current_period_token:
            return c
    return candidate_col


def normalize_company_name(name: str) -> str:
    out = re.sub(r"\s+", " ", name).strip()
    return out


def parse_ta_excel_rows(path: Path, fiscal_year: str, period: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]

    segment_starts: dict[str, int] = {}
    for c in range(1, min(ws.max_column, 40) + 1):
        header = normalize_text(ws.cell(row=4, column=c).value)
        segment = classify_segment(header)
        if segment and segment not in segment_starts:
            segment_starts[segment] = c

    if not segment_starts:
        raise ValueError("Segment headers were not detected in TA workbook.")

    current_period_token = normalize_text(ws.cell(row=5, column=3).value)
    amount_cols: dict[str, int] = {}
    for segment, start_col in segment_starts.items():
        amount_col = find_amount_column(ws, start_col, current_period_token)
        if amount_col is not None:
            amount_cols[segment] = amount_col

    if not amount_cols:
        raise ValueError("Amount columns were not detected in TA workbook.")

    records: list[dict] = []
    for r in range(7, ws.max_row + 1):
        company_raw = normalize_text(ws.cell(row=r, column=2).value)
        if not company_raw:
            continue
        if "会社" in company_raw and "名" in company_raw:
            continue
        if "合計" in company_raw or "参考値" in company_raw:
            continue

        company = normalize_company_name(company_raw)
        for segment, col in amount_cols.items():
            amount = to_float(ws.cell(row=r, column=col).value)
            if amount is None:
                continue
            records.append(
                {
                    "fiscal_year": fiscal_year,
                    "period": period,
                    "company": company,
                    "segment": segment,
                    "amount": amount,
                }
            )

    return records


def build_sqlite(df: pd.DataFrame, sqlite_path: Path) -> None:
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(sqlite_path)) as conn:
        df.to_sql(TABLE_NAME, conn, if_exists="replace", index=False)
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_period "
            f"ON {TABLE_NAME}(fiscal_year, period)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_segment "
            f"ON {TABLE_NAME}(segment)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_company "
            f"ON {TABLE_NAME}(company)"
        )


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    index_html = fetch_html(TA_SOURCE_INDEX_URL)
    fiscal_pages = extract_fiscal_page_links(index_html, TA_SOURCE_INDEX_URL)

    all_records: list[dict] = []
    processed_files: list[dict] = []

    with tempfile.TemporaryDirectory() as td:
        tmp_dir = Path(td)
        file_counter = 0

        for page in fiscal_pages:
            page_url = page["url"]
            page_html = fetch_html(page_url)
            page_soup = BeautifulSoup(page_html, "html.parser")
            page_title = (
                normalize_text(page_soup.title.get_text())
                if page_soup.title is not None
                else ""
            )

            fiscal_year = (
                parse_fiscal_year(page["link_text"])
                or parse_fiscal_year(page_title)
                or ""
            )
            if not fiscal_year:
                print(f"Skipped page (fiscal year unknown): {page_url}")
                continue

            excel_links = extract_ta_excel_links(page_html, page_url)
            for link in excel_links:
                period = parse_period_from_link_text(link["link_text"])
                if period is None:
                    print(f"Skipped link (period unknown): {link['link_text']}")
                    continue

                ext = Path(urlparse(link["url"]).path).suffix.lower() or ".xlsx"
                local_path = tmp_dir / f"ta_{file_counter:04d}{ext}"
                file_counter += 1

                download_file(link["url"], local_path)
                source_sha256 = sha256_file(local_path)

                if ext == ".xls":
                    print(f"Skipped .xls file (unsupported for TA parser): {link['url']}")
                    continue

                try:
                    parsed = parse_ta_excel_rows(
                        path=local_path,
                        fiscal_year=fiscal_year,
                        period=period,
                    )
                except Exception as e:
                    print(f"Skipped (parse failed): {link['url']} ({e})")
                    continue

                if not parsed:
                    print(f"Skipped (no rows): {link['url']}")
                    continue

                all_records.extend(parsed)
                processed_files.append(
                    {
                        "page_url": page_url,
                        "excel_url": link["url"],
                        "excel_link_text": link["link_text"],
                        "sha256": source_sha256,
                        "fiscal_year": fiscal_year,
                        "period": period,
                        "rows": len(parsed),
                    }
                )
                print(
                    f"Parsed TA file: fiscal_year={fiscal_year} period={period} "
                    f"rows={len(parsed)}"
                )

    if not all_records:
        raise RuntimeError("No TA rows were parsed from source pages.")

    df = pd.DataFrame(all_records)
    df["fiscal_year"] = df["fiscal_year"].astype(str)
    df["period"] = df["period"].astype(str)
    df["company"] = df["company"].astype(str)
    df["segment"] = df["segment"].astype(str)
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df.dropna(subset=["amount"]).copy()
    df = df.drop_duplicates(
        subset=["fiscal_year", "period", "company", "segment"], keep="last"
    )

    def period_sort_value(value: str) -> tuple[int, int]:
        if value == "total":
            return 0, 0
        m = re.fullmatch(r"(\d{4})-(\d{2})", value)
        if m:
            return int(m.group(1)), int(m.group(2))
        return 9999, 99

    df["_period_sort_year"] = df["period"].map(lambda p: period_sort_value(p)[0])
    df["_period_sort_month"] = df["period"].map(lambda p: period_sort_value(p)[1])
    df = df.sort_values(
        ["fiscal_year", "_period_sort_year", "_period_sort_month", "segment", "company"]
    ).drop(columns=["_period_sort_year", "_period_sort_month"])
    df = df.reset_index(drop=True)

    build_sqlite(df, SQLITE_PATH)

    meta = {
        "source_index_url": TA_SOURCE_INDEX_URL,
        "last_checked_at": now_utc_iso(),
        "processed_files": processed_files,
        "row_count": int(len(df)),
        "fiscal_years": sorted(df["fiscal_year"].unique().tolist(), reverse=True),
        "periods": sorted(df["period"].unique().tolist()),
        "segments": sorted(df["segment"].unique().tolist()),
        "unit": "thousand_yen",
    }
    save_meta(meta)

    print(
        f"Updated {TABLE_NAME}: rows={len(df)} files={len(processed_files)} "
        f"fiscal_years={len(meta['fiscal_years'])}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
