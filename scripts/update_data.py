from __future__ import annotations

import hashlib
import json
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from .parse_ts_table import (
        build_raw_from_three_sheets,
        parse_facility_occupancy_monthly_sheet,
    )
except ImportError:  # pragma: no cover - fallback for direct script execution
    from parse_ts_table import (
        build_raw_from_three_sheets,
        parse_facility_occupancy_monthly_sheet,
    )

# 取得元（観光庁：宿泊旅行統計調査）
SOURCE_PAGE_URL = "https://www.mlit.go.jp/kankocho/tokei_hakusyo/shukuhakutokei.html"

# 推移表Excel（リンクが固定名のケースに強くする）
PREFERRED_XLSX_NAME_HINT = "001912060.xlsx"

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data"
META_PATH = DATA_DIR / "meta.json"
SQLITE_PATH = DATA_DIR / "market_stats.sqlite"
MARKET_STATS_TABLE_NAME = "market_stats"
STAY_FACILITY_OCCUPANCY_TABLE_NAME = "stay_facility_occupancy"
PIPELINE_VERSION = 4
MONTHLY_SHEET_RE = re.compile(r"^(?P<legacy>旧)?(?P<prefix>[1-8])-(?P<suffix>\d+)$")


def has_required_facility_occupancy_schema(sqlite_path: Path) -> bool:
    import sqlite3

    if not sqlite_path.exists():
        return False
    try:
        with sqlite3.connect(str(sqlite_path)) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (STAY_FACILITY_OCCUPANCY_TABLE_NAME,),
            )
            if cur.fetchone() is None:
                return False
            cur.execute(
                f"PRAGMA table_info({STAY_FACILITY_OCCUPANCY_TABLE_NAME})"
            )
            cols = {str(row[1]) for row in cur.fetchall()}
            required_cols = {
                "ym",
                "pref_code",
                "pref_name",
                "facility_type",
                "occupancy_rate",
            }
            return required_cols.issubset(cols)
    except Exception:
        return False


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_ts_table_xlsx_url(html: str, base_url: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a"):
        href = a.get("href")
        if not href:
            continue
        text = (a.get_text() or "").strip()
        abs_url = urljoin(base_url, href)
        if abs_url.lower().endswith(".xlsx"):
            links.append((abs_url, text))

    if not links:
        raise RuntimeError(
            "No .xlsx links found on source page (HTML structure may have changed)."
        )

    # 1) ファイル名ヒント一致
    for url, _ in links:
        if PREFERRED_XLSX_NAME_HINT in url:
            return url

    # 2) アンカーテキストに「推移表」っぽい語
    for url, text in links:
        if "推移表" in text:
            return url

    # 3) 最初のxlsx（最後の砦）
    return links[0][0]


def fetch_html_text(url: str, timeout_sec: int = 60) -> str:
    response = requests.get(url, timeout=timeout_sec)
    response.raise_for_status()

    # 観光庁ページは requests が ISO-8859-1 と誤判定することがある。
    if (response.encoding or "").lower() == "iso-8859-1":
        apparent = response.apparent_encoding
        if apparent:
            response.encoding = apparent

    return response.text


def find_monthly_sheet_name(workbook, prefix: str, legacy: bool = False) -> str:
    matches = []
    for name in workbook.sheetnames:
        match = MONTHLY_SHEET_RE.match(name)
        if match is None:
            continue
        if match.group("prefix") != prefix:
            continue
        if bool(match.group("legacy")) != legacy:
            continue
        matches.append((int(match.group("suffix")), name))

    if not matches:
        raise KeyError(
            f"Monthly sheet with prefix '{prefix}-' was not found. "
            f"available={workbook.sheetnames}"
        )

    matches.sort(key=lambda item: item[0], reverse=legacy)
    return matches[0][1]


def find_optional_monthly_sheet_name(
    workbook, prefix: str, legacy: bool = False
) -> str | None:
    try:
        return find_monthly_sheet_name(workbook, prefix=prefix, legacy=legacy)
    except KeyError:
        return None


def build_market_stats_from_workbook(workbook) -> pd.DataFrame:
    data_frames = []

    current_total = find_monthly_sheet_name(workbook, "1", legacy=False)
    current_jp = find_monthly_sheet_name(workbook, "2", legacy=False)
    current_foreign = find_monthly_sheet_name(workbook, "3", legacy=False)
    data_frames.append(
        build_raw_from_three_sheets(
            ws_total=workbook[current_total],
            ws_jp=workbook[current_jp],
            ws_foreign=workbook[current_foreign],
            make_national_sum=True,
        )
    )

    legacy_total = find_optional_monthly_sheet_name(workbook, "1", legacy=True)
    legacy_jp = find_optional_monthly_sheet_name(workbook, "2", legacy=True)
    legacy_foreign = find_optional_monthly_sheet_name(workbook, "3", legacy=True)
    if legacy_total and legacy_jp and legacy_foreign:
        data_frames.append(
            build_raw_from_three_sheets(
                ws_total=workbook[legacy_total],
                ws_jp=workbook[legacy_jp],
                ws_foreign=workbook[legacy_foreign],
                make_national_sum=True,
            )
        )

    merged = pd.concat(data_frames, ignore_index=True)
    merged = merged.drop_duplicates(subset=["ym", "pref_code"], keep="first")
    return merged.sort_values(["ym", "pref_code"]).reset_index(drop=True)


def build_facility_occupancy_from_workbook(workbook) -> pd.DataFrame:
    data_frames = []

    current_sheet = find_monthly_sheet_name(workbook, "4", legacy=False)
    data_frames.append(
        parse_facility_occupancy_monthly_sheet(workbook[current_sheet])
    )

    legacy_sheet = find_optional_monthly_sheet_name(workbook, "4", legacy=True)
    if legacy_sheet:
        data_frames.append(
            parse_facility_occupancy_monthly_sheet(workbook[legacy_sheet])
        )

    merged = pd.concat(data_frames, ignore_index=True)
    merged = merged.drop_duplicates(
        subset=["ym", "pref_code", "facility_type"], keep="first"
    )
    return merged.sort_values(["ym", "pref_code", "facility_type"]).reset_index(
        drop=True
    )


def load_meta() -> dict:
    if not META_PATH.exists():
        return {}
    return json.loads(META_PATH.read_text(encoding="utf-8"))


def save_meta(meta: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    META_PATH.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def download_file(url: str, dst: Path, timeout_sec: int = 60) -> None:
    with requests.get(url, stream=True, timeout=timeout_sec) as r:
        r.raise_for_status()
        with dst.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 64):
                if chunk:
                    f.write(chunk)


def build_sqlite(
    df_market_stats: pd.DataFrame,
    df_stay_facility_occupancy: pd.DataFrame,
    sqlite_path: Path,
) -> None:
    import sqlite3

    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(sqlite_path)) as conn:
        df_market_stats.to_sql(
            MARKET_STATS_TABLE_NAME, conn, if_exists="replace", index=False
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_market_stats_ym ON {MARKET_STATS_TABLE_NAME}(ym)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_market_stats_pref ON {MARKET_STATS_TABLE_NAME}(pref_code)"
        )
        df_stay_facility_occupancy.to_sql(
            STAY_FACILITY_OCCUPANCY_TABLE_NAME, conn, if_exists="replace", index=False
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_stay_facility_occupancy_ym "
            f"ON {STAY_FACILITY_OCCUPANCY_TABLE_NAME}(ym)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_stay_facility_occupancy_pref "
            f"ON {STAY_FACILITY_OCCUPANCY_TABLE_NAME}(pref_code)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_stay_facility_occupancy_type "
            f"ON {STAY_FACILITY_OCCUPANCY_TABLE_NAME}(facility_type)"
        )


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    html = fetch_html_text(SOURCE_PAGE_URL, timeout_sec=60)
    xlsx_url = find_ts_table_xlsx_url(html, SOURCE_PAGE_URL)

    with tempfile.TemporaryDirectory() as td:
        tmp_xlsx = Path(td) / "ts_table.xlsx"
        download_file(xlsx_url, tmp_xlsx)

        fetched_sha = sha256_file(tmp_xlsx)
        meta = load_meta()
        if (
            meta.get("source_sha256") == fetched_sha
            and int(meta.get("pipeline_version", 0)) == PIPELINE_VERSION
            and has_required_facility_occupancy_schema(SQLITE_PATH)
        ):
            print("No change: source file hash unchanged.")
            return 0

        wb = load_workbook(tmp_xlsx, read_only=False, data_only=True)
        try:
            df = build_market_stats_from_workbook(wb)
            df_facility_occupancy = build_facility_occupancy_from_workbook(wb)
        finally:
            wb.close()

        # 型整形
        df["ym"] = df["ym"].astype(str)
        df["pref_code"] = df["pref_code"].astype(str)
        df["pref_name"] = df["pref_name"].astype(str)
        df_facility_occupancy["ym"] = df_facility_occupancy["ym"].astype(str)
        df_facility_occupancy["pref_code"] = (
            df_facility_occupancy["pref_code"].astype(str).str.zfill(2)
        )
        df_facility_occupancy["pref_name"] = df_facility_occupancy["pref_name"].astype(
            str
        )
        df_facility_occupancy["facility_type"] = df_facility_occupancy[
            "facility_type"
        ].astype(str)
        df_facility_occupancy["occupancy_rate"] = pd.to_numeric(
            df_facility_occupancy["occupancy_rate"], errors="coerce"
        )
        df_facility_occupancy = df_facility_occupancy.dropna(
            subset=["occupancy_rate"]
        ).copy()

        build_sqlite(df, df_facility_occupancy, SQLITE_PATH)

        now = datetime.now(timezone.utc).isoformat()
        new_meta = {
            "source_page_url": SOURCE_PAGE_URL,
            "source_xlsx_url": xlsx_url,
            "source_sha256": fetched_sha,
            "pipeline_version": PIPELINE_VERSION,
            "fetched_at_utc": now,
            "rows": int(len(df)),
            "min_ym": str(df["ym"].min()),
            "max_ym": str(df["ym"].max()),
            "facility_occupancy_rows": int(len(df_facility_occupancy)),
            "facility_occupancy_min_ym": str(df_facility_occupancy["ym"].min()),
            "facility_occupancy_max_ym": str(df_facility_occupancy["ym"].max()),
        }
        save_meta(new_meta)

        print(
            "Updated: "
            f"rows={len(df)} ym={new_meta['min_ym']}..{new_meta['max_ym']} "
            f"facility_rows={len(df_facility_occupancy)} "
            f"facility_ym={new_meta['facility_occupancy_min_ym']}..{new_meta['facility_occupancy_max_ym']}"
        )
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
