from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class HeaderLayout:
    year_row: int
    month_row: int
    first_data_col: int


MONTH_RE = re.compile(r"^\s*(\d{1,2})\s*月\s*$")
ERA_RE = re.compile(r"^\s*(昭和|平成|令和)\s*([0-9]{1,2}|元)\s*年\s*$")


def _to_int_era_year(s: str) -> Tuple[str, int]:
    """
    '平成23年' -> ('平成', 23)
    '令和元年' -> ('令和', 1)
    """
    m = ERA_RE.match(str(s))
    if not m:
        raise ValueError(f"Unsupported era-year label: {s!r}")
    era = m.group(1)
    y = m.group(2)
    n = 1 if y == "元" else int(y)
    return era, n


def _era_to_gregorian(era: str, year: int) -> int:
    # Showa 1 = 1926, Heisei 1 = 1989, Reiwa 1 = 2019
    if era == "昭和":
        return 1925 + year
    if era == "平成":
        return 1988 + year
    if era == "令和":
        return 2018 + year
    raise ValueError(f"Unknown era: {era}")


def detect_layout(
    ws: Worksheet, scan_rows: int = 30, scan_cols: int = 80
) -> HeaderLayout:
    """
    推移表（例: 1-2 / 2-2 / 3-2）想定。
    - month_row: '1月' が並ぶ行
    - year_row: その1行上（'平成xx年' / '令和x年' が入る）
    """
    month_row = None
    first_data_col = None

    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if MONTH_RE.match(str(v)):
                month_row = r
                first_data_col = c
                break
        if month_row is not None:
            break

    if month_row is None or first_data_col is None:
        raise ValueError("Failed to detect month header row/col (e.g., '1月').")

    year_row = month_row - 1
    return HeaderLayout(
        year_row=year_row, month_row=month_row, first_data_col=first_data_col
    )


def build_ym_by_col(
    ws: Worksheet, layout: HeaderLayout, max_cols: int = 200
) -> List[Tuple[int, str]]:
    """
    Returns: list of (col_index, ym='YYYY-MM') for each month column.
    """
    yms: List[Tuple[int, str]] = []
    current_year_label: Optional[str] = None

    for c in range(layout.first_data_col, layout.first_data_col + max_cols):
        month_v = ws.cell(layout.month_row, c).value
        if month_v is None:
            # ヘッダ終端（連続None）を許容
            if len(yms) > 0:
                # 次以降もNoneが続くなら打ち切り
                next_v = ws.cell(layout.month_row, c + 1).value
                if next_v is None:
                    break
            continue

        m = MONTH_RE.match(str(month_v))
        if not m:
            # 月表示でない＝終端
            if len(yms) > 0:
                break
            continue

        # 年ラベル（結合セル）を左→右に伝播
        year_v = ws.cell(layout.year_row, c).value
        if year_v is not None:
            current_year_label = str(year_v).strip()

        if not current_year_label:
            raise ValueError(f"Missing year label around col={c} (month={month_v}).")

        era, era_year = _to_int_era_year(current_year_label)
        year_g = _era_to_gregorian(era, era_year)
        month_i = int(m.group(1))
        ym = f"{year_g:04d}-{month_i:02d}"
        yms.append((c, ym))

    if not yms:
        raise ValueError("No YM columns detected.")
    return yms


def iter_pref_rows(
    ws: Worksheet, start_row: int, max_rows: int = 200
) -> Iterable[Tuple[int, str, str]]:
    """
    Returns rows: (row_index, pref_code, pref_name)
    - '01北海道' -> ('01', '北海道')
    - '全 国' / '全国' -> ('00', '全国')  ※ただしMVPでは合算生成するので後で除外推奨
    """
    for r in range(start_row, start_row + max_rows):
        v = ws.cell(r, 1).value
        if v is None:
            # 連続空で打ち切り
            nxt = ws.cell(r + 1, 1).value
            if nxt is None:
                break
            continue

        s = str(v).strip().replace("　", "").replace(" ", "")
        if not s:
            continue

        if s.startswith("全"):
            yield r, "00", "全国"
            continue

        m = re.match(r"^(\d{2})(.+)$", s)
        if not m:
            # 想定外行はスキップ（注記行など）
            continue
        yield r, m.group(1), m.group(2)


def parse_sheet_long(ws: Worksheet, metric: str) -> pd.DataFrame:
    """
    metric in {'total','jp','foreign'}
    Output columns: ym, pref_code, pref_name, metric, value
    """
    layout = detect_layout(ws)
    ym_cols = build_ym_by_col(ws, layout)

    # pref starts the row right below month header
    pref_start_row = layout.month_row + 1
    rows = []
    for r, pref_code, pref_name in iter_pref_rows(ws, start_row=pref_start_row):
        for c, ym in ym_cols:
            val = ws.cell(r, c).value
            if val is None:
                continue
            if isinstance(val, (int, float)):
                rows.append(
                    {
                        "ym": ym,
                        "pref_code": pref_code,
                        "pref_name": pref_name,
                        "metric": metric,
                        "value": float(val),
                    }
                )

    if not rows:
        raise ValueError(f"No numeric rows parsed for metric={metric}.")
    return pd.DataFrame(rows)


def _normalize_compact_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\n", "")
    text = re.sub(r"\s+", "", text)
    return text.strip()


def _to_float_or_none(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if text in {"", "-", "－", "―", "‐", "…"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_pref_cell(value: object) -> Optional[Tuple[str, str]]:
    s = _normalize_compact_text(value)
    if not s:
        return None
    if s.startswith("全"):
        return "00", "全国"
    m = re.match(r"^(\d{2})(.+)$", s)
    if not m:
        return None
    return m.group(1), m.group(2)


def parse_facility_occupancy_monthly_sheet(
    ws: Worksheet, max_rows: int = 500, max_cols: int = 200
) -> pd.DataFrame:
    """
    Parse sheet 4-2 style monthly occupancy table (national + prefectures).

    Output columns: ym, pref_code, pref_name, facility_type, occupancy_rate
    """
    layout = detect_layout(ws)
    ym_cols = build_ym_by_col(ws, layout, max_cols=max_cols)

    pref_start_row = layout.month_row + 1
    rows = []
    current_pref_code = ""
    current_pref_name = ""
    blank_streak = 0

    for r in range(pref_start_row, pref_start_row + max_rows):
        pref_raw = ws.cell(r, 1).value
        facility_raw = ws.cell(r, 2).value

        if pref_raw is None and facility_raw is None:
            blank_streak += 1
            if blank_streak >= 6 and rows:
                break
            continue
        blank_streak = 0

        parsed_pref = _parse_pref_cell(pref_raw)
        if parsed_pref is not None:
            current_pref_code, current_pref_name = parsed_pref

        facility_type = _normalize_compact_text(facility_raw)
        if not facility_type:
            continue
        if not current_pref_code or not current_pref_name:
            continue

        for c, ym in ym_cols:
            value = _to_float_or_none(ws.cell(r, c).value)
            if value is None:
                continue
            rows.append(
                {
                    "ym": ym,
                    "pref_code": current_pref_code,
                    "pref_name": current_pref_name,
                    "facility_type": facility_type,
                    "occupancy_rate": value,
                }
            )

    if not rows:
        raise ValueError("No facility occupancy rows were parsed.")

    out = pd.DataFrame(rows)
    out = out.drop_duplicates(subset=["ym", "pref_code", "facility_type"])
    out = out.sort_values(["ym", "pref_code", "facility_type"]).reset_index(drop=True)
    return out


def parse_national_facility_occupancy_monthly_sheet(
    ws: Worksheet, max_rows: int = 500, max_cols: int = 200
) -> pd.DataFrame:
    out = parse_facility_occupancy_monthly_sheet(
        ws, max_rows=max_rows, max_cols=max_cols
    )
    out = out[out["pref_code"] == "00"].copy()
    out = out[["ym", "facility_type", "occupancy_rate"]]
    return out.reset_index(drop=True)


def build_raw_from_three_sheets(
    ws_total: Worksheet,
    ws_jp: Worksheet,
    ws_foreign: Worksheet,
    make_national_sum: bool = True,
) -> pd.DataFrame:
    """
    Returns wide RAW:
      ym, pref_code, pref_name, total, jp, foreign
    """
    df_t = parse_sheet_long(ws_total, "total")
    df_j = parse_sheet_long(ws_jp, "jp")
    df_f = parse_sheet_long(ws_foreign, "foreign")

    df = pd.concat([df_t, df_j, df_f], ignore_index=True)

    wide = df.pivot_table(
        index=["ym", "pref_code", "pref_name"],
        columns="metric",
        values="value",
        aggfunc="sum",
    ).reset_index()

    # 欠損を0に（シート構造差の保険）
    for col in ["total", "jp", "foreign"]:
        if col not in wide.columns:
            wide[col] = 0.0
        wide[col] = wide[col].fillna(0.0)

    # 全国行はファイル由来を採用せず、都道府県合算で生成（ズレ耐性）
    if make_national_sum:
        base = wide[wide["pref_code"] != "00"].copy()
        nat = base.groupby(["ym"], as_index=False)[["total", "jp", "foreign"]].sum()
        nat["pref_code"] = "00"
        nat["pref_name"] = "全国"
        wide = pd.concat([base, nat[wide.columns]], ignore_index=True)

    # ym sort
    wide = wide.sort_values(["ym", "pref_code"]).reset_index(drop=True)
    return wide
