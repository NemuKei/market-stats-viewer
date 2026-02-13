from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 観光庁: 旅行・観光消費動向調査
TCD_SOURCE_PAGE_URL = "https://www.mlit.go.jp/kankocho/siryou/toukei/shouhidoukou.html"

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data"
SQLITE_PATH = DATA_DIR / "market_stats.sqlite"
META_TCD_PATH = DATA_DIR / "meta_tcd.json"
TABLE_NAME = "tcd_stay_nights"

NIGHTS_BIN_ORDER = ["1泊", "2泊", "3泊", "4泊", "5泊", "6泊", "7泊", "8泊以上"]
TARGET_SEGMENTS = [("domestic_total", 2), ("domestic_business", 5)]


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def load_meta_tcd() -> dict:
    if not META_TCD_PATH.exists():
        return {}
    return json.loads(META_TCD_PATH.read_text(encoding="utf-8"))


def save_meta_tcd(meta: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    META_TCD_PATH.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def fetch_source_page(url: str) -> str:
    res = requests.get(url, timeout=60)
    res.raise_for_status()
    if not res.encoding or res.encoding.lower() == "iso-8859-1":
        res.encoding = res.apparent_encoding
    return res.text


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def extract_target_excel_links(html: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[dict[str, str]] = []
    seen: set[str] = set()

    for a in soup.find_all("a"):
        href = normalize_text(a.get("href"))
        text = normalize_text(a.get_text())
        if not href:
            continue

        abs_url = urljoin(base_url, href)
        lower_url = abs_url.lower()
        if not (lower_url.endswith(".xlsx") or lower_url.endswith(".xls")):
            continue

        # サイト本文の文字化け・表記揺れに備え、URLヒントでも拾う。
        compact_text = text.replace(" ", "").replace("　", "")
        has_target_hint = ("集計表" in compact_text) or ("/content/" in lower_url)
        if not has_target_hint:
            continue

        # 明確に対象外と判定できる場合のみ除外。
        if "都道府県" in compact_text and "参考" in compact_text:
            continue

        if abs_url in seen:
            continue
        seen.add(abs_url)
        out.append({"url": abs_url, "link_text": text})

    if not out:
        raise RuntimeError("Target Excel links were not found on source page.")

    return out


def download_file(url: str, dst: Path) -> None:
    with requests.get(url, stream=True, timeout=120) as res:
        res.raise_for_status()
        with dst.open("wb") as f:
            for chunk in res.iter_content(chunk_size=1024 * 64):
                if chunk:
                    f.write(chunk)


def get_title_a1(workbook: Workbook) -> str:
    if "表題" in workbook.sheetnames:
        ws = workbook["表題"]
    else:
        ws = workbook[workbook.sheetnames[0]]
    return normalize_text(ws["A1"].value)


def normalize_release_type(text: str) -> str | None:
    if "確報" in text:
        return "確報"
    if "2次速報" in text or "２次速報" in text:
        return "2次速報"
    return None


def parse_period_from_text(text: str) -> tuple[str, str, str] | None:
    normalized = (
        text.replace("〜", "-")
        .replace("～", "-")
        .replace("−", "-")
        .replace("－", "-")
        .replace("―", "-")
    )

    # 例: 2025年1-3月期
    m_quarter_range = re.search(
        r"(20\d{2})年\s*([1-9]|1[0-2])\s*-\s*([1-9]|1[0-2])月",
        normalized,
    )
    if m_quarter_range:
        year = int(m_quarter_range.group(1))
        start_month = int(m_quarter_range.group(2))
        quarter = (start_month - 1) // 3 + 1
        key = f"{year}Q{quarter}"
        return "quarter", key, f"{year}年Q{quarter}"

    # 例: 2025年Q1
    m_quarter_q = re.search(r"(20\d{2})年\s*Q([1-4])", normalized, flags=re.IGNORECASE)
    if m_quarter_q:
        year = int(m_quarter_q.group(1))
        quarter = int(m_quarter_q.group(2))
        key = f"{year}Q{quarter}"
        return "quarter", key, f"{year}年Q{quarter}"

    # 例: 2024年
    m_annual = re.search(r"(20\d{2})年", normalized)
    if m_annual:
        year = m_annual.group(1)
        return "annual", year, f"{year}年"

    return None


def parse_title_metadata(title_a1: str, link_text: str) -> tuple[str, str, str, str]:
    period = parse_period_from_text(title_a1)
    if period is None:
        raise ValueError(f"Failed to parse period from title A1: {title_a1}")
    period_type, period_key, period_label = period

    release_type = normalize_release_type(title_a1) or normalize_release_type(link_text)
    if release_type not in {"確報", "2次速報"}:
        raise ValueError(f"Unsupported release type in title/link: {title_a1} / {link_text}")

    return period_type, period_key, period_label, release_type


def normalize_nights_bin(value: object) -> str | None:
    s = normalize_text(value).replace(" ", "").replace("　", "")
    if not s:
        return None

    if "8泊" in s and "以上" in s:
        return "8泊以上"

    m = re.match(r"^([1-7])泊", s)
    if m:
        return f"{m.group(1)}泊"

    if s in NIGHTS_BIN_ORDER:
        return s
    return None


def to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = normalize_text(value).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_t06_section_rows(ws: Worksheet) -> list[int]:
    starts: list[int] = []
    for row in range(1, ws.max_row + 1):
        a = normalize_text(ws.cell(row=row, column=1).value).replace(" ", "").replace("　", "")
        if a == "宿泊数":
            starts.append(row)
    return starts


def find_period_for_section(
    ws: Worksheet,
    section_row: int,
    fallback: tuple[str, str, str],
) -> tuple[str, str, str]:
    for row in range(section_row - 1, max(section_row - 20, 0), -1):
        for col in range(1, 7):
            text = normalize_text(ws.cell(row=row, column=col).value)
            if not text:
                continue
            parsed = parse_period_from_text(text)
            if parsed is not None:
                return parsed
    return fallback


def extract_t06_rows(
    workbook: Workbook,
    source_url: str,
    source_title: str,
    source_sha256: str,
    title_period_fallback: tuple[str, str, str],
    release_type: str,
) -> pd.DataFrame:
    if "T06" not in workbook.sheetnames:
        raise ValueError("Sheet T06 not found.")
    ws = workbook["T06"]

    section_rows = find_t06_section_rows(ws)
    if not section_rows:
        raise ValueError("Section key '宿泊数' was not found in T06.")

    records: list[dict] = []
    for section_row in section_rows:
        period_type, period_key, period_label = find_period_for_section(
            ws, section_row, title_period_fallback
        )

        for offset in range(1, 9):
            row = section_row + offset
            nights_bin = normalize_nights_bin(ws.cell(row=row, column=1).value)
            if nights_bin is None:
                continue

            for segment, col in TARGET_SEGMENTS:
                value = to_float(ws.cell(row=row, column=col).value)
                if value is None:
                    continue
                records.append(
                    {
                        "period_type": period_type,
                        "period_key": period_key,
                        "period_label": period_label,
                        "release_type": release_type,
                        "segment": segment,
                        "nights_bin": nights_bin,
                        "value": value,
                        "source_url": source_url,
                        "source_title": source_title,
                        "source_sha256": source_sha256,
                    }
                )

    if not records:
        raise ValueError("No rows parsed from T06 sections.")

    df = pd.DataFrame(records)
    return df


def load_existing_tcd_rows(sqlite_path: Path) -> pd.DataFrame:
    if not sqlite_path.exists():
        return pd.DataFrame()

    with sqlite3.connect(str(sqlite_path)) as conn:
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (TABLE_NAME,),
        ).fetchone()
        if not table_exists:
            return pd.DataFrame()
        return pd.read_sql_query(f"SELECT * FROM {TABLE_NAME}", conn)


def build_tcd_sqlite(df: pd.DataFrame, sqlite_path: Path) -> None:
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(sqlite_path)) as conn:
        df.to_sql(TABLE_NAME, conn, if_exists="replace", index=False)
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_period ON {TABLE_NAME}(period_type, period_key)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_release ON {TABLE_NAME}(release_type)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_source ON {TABLE_NAME}(source_url, source_sha256)"
        )


def build_available_periods(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []

    records: list[dict] = []
    grouped = df.groupby(["period_type", "period_key", "period_label"], dropna=False)
    for (period_type, period_key, period_label), g in grouped:
        releases = sorted(
            g["release_type"].dropna().astype(str).unique().tolist(),
            key=lambda x: 0 if x == "確報" else 1,
        )
        records.append(
            {
                "period_type": str(period_type),
                "period_key": str(period_key),
                "period_label": str(period_label),
                "releases": releases,
            }
        )

    def sort_key(item: dict) -> tuple[int, int]:
        key = item["period_key"]
        m_quarter = re.fullmatch(r"(\d{4})Q([1-4])", key)
        if m_quarter:
            return int(m_quarter.group(1)), int(m_quarter.group(2))
        m_annual = re.fullmatch(r"(\d{4})", key)
        if m_annual:
            return int(m_annual.group(1)), 0
        return 0, 0

    return sorted(records, key=sort_key, reverse=True)


def empty_tcd_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "period_type",
            "period_key",
            "period_label",
            "release_type",
            "segment",
            "nights_bin",
            "value",
            "source_url",
            "source_title",
            "source_sha256",
        ]
    )


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    html = fetch_source_page(TCD_SOURCE_PAGE_URL)
    links = extract_target_excel_links(html, TCD_SOURCE_PAGE_URL)

    old_meta = load_meta_tcd()
    old_processed = {
        str(x.get("url")): str(x.get("sha256"))
        for x in old_meta.get("processed_files", [])
        if x.get("url") and x.get("sha256")
    }
    old_titles = {
        str(x.get("url")): str(x.get("title_a1"))
        for x in old_meta.get("processed_files", [])
        if x.get("url")
    }
    old_rows = load_existing_tcd_rows(SQLITE_PATH)

    fetched_files: list[dict[str, str]] = []
    with tempfile.TemporaryDirectory() as td:
        for idx, link in enumerate(links):
            url = link["url"]
            local_path = Path(td) / f"tcd_{idx:03d}.xlsx"
            download_file(url, local_path)
            fetched_files.append(
                {
                    "url": url,
                    "link_text": link["link_text"],
                    "local_path": str(local_path),
                    "sha256": sha256_file(local_path),
                }
            )

        current_processed = {x["url"]: x["sha256"] for x in fetched_files}
        source_set_changed = set(current_processed.keys()) != set(old_processed.keys())
        hash_changed = any(old_processed.get(url) != sha for url, sha in current_processed.items())
        old_rows_available = not old_rows.empty

        if not source_set_changed and not hash_changed and old_rows_available:
            print("No change: source file hash set unchanged.")
            return 0

        rebuilt_parts: list[pd.DataFrame] = []
        fetched_at = now_utc_iso()
        processed_entries: list[dict] = []

        has_reuse_columns = {"source_url", "source_sha256"}.issubset(old_rows.columns)
        for item in fetched_files:
            url = item["url"]
            sha = item["sha256"]
            link_text = item["link_text"]
            local_path = Path(item["local_path"])
            title_a1 = old_titles.get(url, "")

            can_reuse = (
                old_processed.get(url) == sha
                and has_reuse_columns
                and not old_rows.empty
            )
            reused = pd.DataFrame()
            if can_reuse:
                reused = old_rows[
                    (old_rows["source_url"] == url) & (old_rows["source_sha256"] == sha)
                ].copy()

            if not reused.empty:
                title_a1 = normalize_text(reused["source_title"].iloc[0])
                rebuilt_parts.append(reused)
                processed_entries.append(
                    {
                        "url": url,
                        "sha256": sha,
                        "title_a1": title_a1,
                        "fetched_at": fetched_at,
                    }
                )
                print(f"Reused cached rows: {url}")
                continue

            try:
                wb = load_workbook(local_path, read_only=False, data_only=True)
                title_a1 = get_title_a1(wb)
            except Exception as e:
                processed_entries.append(
                    {
                        "url": url,
                        "sha256": sha,
                        "title_a1": title_a1,
                        "fetched_at": fetched_at,
                    }
                )
                print(f"Skipped (open failed): {url} ({e})")
                continue

            processed_entries.append(
                {
                    "url": url,
                    "sha256": sha,
                    "title_a1": title_a1,
                    "fetched_at": fetched_at,
                }
            )

            try:
                period_type, period_key, period_label, release_type = parse_title_metadata(
                    title_a1, link_text
                )
                parsed = extract_t06_rows(
                    workbook=wb,
                    source_url=url,
                    source_title=title_a1,
                    source_sha256=sha,
                    title_period_fallback=(period_type, period_key, period_label),
                    release_type=release_type,
                )
            except Exception as e:
                print(f"Skipped (non-target/unsupported): {url} ({e})")
                continue

            rebuilt_parts.append(parsed)
            print(f"Parsed: {url}")

        if rebuilt_parts:
            new_df = pd.concat(rebuilt_parts, ignore_index=True)
        else:
            raise RuntimeError(
                "No parsable TCD files found after download. Check source page structure."
            )

        if not new_df.empty:
            new_df["period_key"] = new_df["period_key"].astype(str)
            new_df["nights_bin"] = new_df["nights_bin"].astype(str)
            new_df["segment"] = new_df["segment"].astype(str)
            new_df["release_type"] = new_df["release_type"].astype(str)
            new_df["period_type"] = new_df["period_type"].astype(str)
            new_df["value"] = pd.to_numeric(new_df["value"], errors="coerce").fillna(0.0)

            nights_order = {label: i for i, label in enumerate(NIGHTS_BIN_ORDER)}
            new_df["_nights_sort"] = new_df["nights_bin"].map(lambda x: nights_order.get(x, 999))
            new_df = new_df.sort_values(
                [
                    "period_type",
                    "period_key",
                    "release_type",
                    "_nights_sort",
                    "segment",
                    "source_url",
                ]
            ).drop(columns=["_nights_sort"])
            new_df = new_df.reset_index(drop=True)

        build_tcd_sqlite(new_df, SQLITE_PATH)

        new_meta = {
            "source_page_url": TCD_SOURCE_PAGE_URL,
            "last_checked_at": fetched_at,
            "processed_files": processed_entries,
            "available_periods": build_available_periods(new_df),
            "note": "最新（確報優先）は同一period_type内で最新period_keyを選び、同一期に確報と2次速報がある場合は確報を優先する。",
        }
        save_meta_tcd(new_meta)

        print(
            f"Updated {TABLE_NAME}: rows={len(new_df)} files={len(processed_entries)} "
            f"periods={len(new_meta['available_periods'])}"
        )
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
