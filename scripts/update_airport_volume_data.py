from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

SOURCE_NAME = "e-Stat \u51fa\u5165\u56fd\u7ba1\u7406\u7d71\u8a08\uff08\u6e2f\u5225 \u51fa\u5165\u56fd\u8005\uff09"
SOURCE_LIST_BASE_URL = (
    "https://www.e-stat.go.jp/stat-search/files"
    "?cycle=1&cycle_facet=tclass1%3Acycle&layout=dataset&result_back=1"
    "&tclass1=000001012481&tclass2val=0&toukei=00250011&tstat=000001012480"
)
TARGET_TITLE_RE = re.compile(
    r"^\u7dcf\u62ec (\d{2})-(\d{2})-01 \u6e2f\u5225 \u51fa\u5165\u56fd\u8005$"
)

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data"
SQLITE_PATH = DATA_DIR / "market_stats.sqlite"
META_PATH = DATA_DIR / "meta_airport_volume.json"
TABLE_NAME = "airport_arrivals_monthly"
UNIT = "persons"
MAX_MONTHS = 48
MAX_PAGES = 12

# MVP: ICD entry_port との将来突合を見据えて主要空港のみ保持する。
AIRPORT_MAPPING: dict[str, tuple[str, str]] = {
    "\u65b0\u5343\u6b73\uff08\u7a7a\u6e2f\uff09": ("\u65b0\u5343\u6b73\u7a7a\u6e2f", "CTS"),
    "\u6210\u7530\uff08\u7a7a\u6e2f\uff09": ("\u6210\u7530\u56fd\u969b\u7a7a\u6e2f", "NRT"),
    "\u7fbd\u7530\uff08\u7a7a\u6e2f\uff09": ("\u6771\u4eac\u56fd\u969b\u7a7a\u6e2f\uff08\u7fbd\u7530\uff09", "HND"),
    "\u4e2d\u90e8\uff08\u7a7a\u6e2f\uff09": ("\u4e2d\u90e8\u56fd\u969b\u7a7a\u6e2f", "NGO"),
    "\u95a2\u897f\uff08\u7a7a\u6e2f\uff09": ("\u95a2\u897f\u56fd\u969b\u7a7a\u6e2f", "KIX"),
    "\u798f\u5ca1\uff08\u7a7a\u6e2f\uff09": ("\u798f\u5ca1\u7a7a\u6e2f", "FUK"),
    "\u90a3\u8987\uff08\u7a7a\u6e2f\uff09": ("\u90a3\u8987\u7a7a\u6e2f", "OKA"),
    "\u4ed9\u53f0\uff08\u7a7a\u6e2f\uff09": ("\u4ed9\u53f0\u7a7a\u6e2f", "SDJ"),
    "\u65b0\u6f5f\uff08\u7a7a\u6e2f\uff09": ("\u65b0\u6f5f\u7a7a\u6e2f", "KIJ"),
    "\u9e7f\u5150\u5cf6\uff08\u7a7a\u6e2f\uff09": ("\u9e7f\u5150\u5cf6\u7a7a\u6e2f", "KOJ"),
    "\u5e83\u5cf6\uff08\u7a7a\u6e2f\uff09": ("\u5e83\u5cf6\u7a7a\u6e2f", "HIJ"),
    "\u718a\u672c\uff08\u7a7a\u6e2f\uff09": ("\u718a\u672c\u7a7a\u6e2f", "KMJ"),
    "\u9577\u5d0e\uff08\u7a7a\u6e2f\uff09": ("\u9577\u5d0e\u7a7a\u6e2f", "NGS"),
}


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def build_source_list_url(page: int) -> str:
    return f"{SOURCE_LIST_BASE_URL}&page={page}"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def load_meta() -> dict:
    if not META_PATH.exists():
        return {}
    return json.loads(META_PATH.read_text(encoding="utf-8"))


def save_meta(meta: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    META_PATH.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_html(url: str) -> str:
    res = requests.get(url, timeout=60)
    res.raise_for_status()
    if not res.encoding or res.encoding.lower() == "iso-8859-1":
        res.encoding = res.apparent_encoding
    return res.text


def download_file(url: str, dst: Path) -> None:
    with requests.get(url, stream=True, timeout=120) as res:
        res.raise_for_status()
        with dst.open("wb") as f:
            for chunk in res.iter_content(chunk_size=1024 * 64):
                if chunk:
                    f.write(chunk)


def extract_source_entries(html: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")

    entries: list[dict[str, str]] = []
    for article in soup.select("article.stat-resource_list-item"):
        article_text = normalize_text(article.get_text(" ", strip=True))

        title_link = None
        title_text = ""
        for a in article.find_all("a", href=True):
            text = normalize_text(a.get_text())
            if TARGET_TITLE_RE.fullmatch(text):
                title_link = a
                title_text = text
                break
        if title_link is None:
            continue

        m_title = TARGET_TITLE_RE.fullmatch(title_text)
        if m_title is None:
            continue
        year = 2000 + int(m_title.group(1))
        month = int(m_title.group(2))
        period_key = f"{year:04d}-{month:02d}"

        excel_url = ""
        for a in article.find_all("a", href=True):
            text = normalize_text(a.get_text())
            if "EXCEL" in text and "\u95b2\u89a7\u7528" in text:
                excel_url = urljoin(base_url, a["href"])
                break
        if not excel_url:
            continue

        release_date = ""
        m_release = re.search(
            r"\u516c\u958b\uff08\u66f4\u65b0\uff09\u65e5 (\d{4}-\d{2}-\d{2})",
            article_text,
        )
        if m_release:
            release_date = m_release.group(1)

        entries.append(
            {
                "title": title_text,
                "detail_url": urljoin(base_url, title_link["href"]),
                "excel_url": excel_url,
                "period_key": period_key,
                "release_date": release_date,
            }
        )

    return entries


def collect_recent_source_entries(max_months: int = MAX_MONTHS) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen_periods: set[str] = set()

    for page in range(1, MAX_PAGES + 1):
        list_url = build_source_list_url(page)
        html = fetch_html(list_url)
        entries = extract_source_entries(html, list_url)
        if not entries:
            break

        for entry in entries:
            period_key = entry["period_key"]
            if period_key in seen_periods:
                continue
            seen_periods.add(period_key)
            out.append(entry)
            if len(out) >= max_months:
                return out

    return out


def detect_period_key_from_sheet(path: Path, fallback: str) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        maybe_date = ws.cell(row=2, column=6).value
        if isinstance(maybe_date, datetime):
            return f"{maybe_date.year:04d}-{maybe_date.month:02d}"

        maybe_text = normalize_text(maybe_date)
        m = re.search(r"(20\d{2})[^\d]?([0-9]{1,2})", maybe_text)
        if m:
            return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

        return fallback
    finally:
        wb.close()


def parse_airport_rows(
    path: Path,
    period_key: str,
    source_url: str,
    updated_at_utc: str,
) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        records: list[dict] = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            airport_name_raw = normalize_text(row[1] if len(row) > 1 else "")
            arrivals_raw = row[2] if len(row) > 2 else None

            if not airport_name_raw:
                continue
            mapped = AIRPORT_MAPPING.get(airport_name_raw)
            if mapped is None:
                continue

            try:
                arrivals = int(float(arrivals_raw)) if arrivals_raw is not None else 0
            except (TypeError, ValueError):
                continue

            airport_name, airport_code = mapped
            records.append(
                {
                    "period_key": period_key,
                    "airport_name_raw": airport_name_raw,
                    "airport_name": airport_name,
                    "airport_code": airport_code,
                    "arrivals": arrivals,
                    "unit": UNIT,
                    "source_name": SOURCE_NAME,
                    "source_url": source_url,
                    "updated_at_utc": updated_at_utc,
                }
            )
    finally:
        wb.close()

    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def build_source_signature(downloaded: list[dict[str, str]]) -> str:
    payload = "\n".join(
        f"{item['period_key']}|{item['excel_url']}|{item['sha256']}"
        for item in sorted(downloaded, key=lambda x: x["period_key"])
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_sqlite(df: pd.DataFrame, sqlite_path: Path) -> None:
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(sqlite_path)) as conn:
        df.to_sql(TABLE_NAME, conn, if_exists="replace", index=False)
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_period "
            f"ON {TABLE_NAME}(period_key)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_airport "
            f"ON {TABLE_NAME}(airport_code)"
        )


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    sources = collect_recent_source_entries(MAX_MONTHS)
    if not sources:
        raise RuntimeError("No source entries were found in e-Stat dataset listing.")

    old_meta = load_meta()

    with tempfile.TemporaryDirectory() as td:
        tmp_dir = Path(td)
        downloaded: list[dict[str, str]] = []
        for i, source in enumerate(sources):
            local_path = tmp_dir / f"airport_arrivals_{i:03d}.xlsx"
            download_file(source["excel_url"], local_path)
            downloaded.append(
                {
                    "period_key": source["period_key"],
                    "title": source["title"],
                    "detail_url": source["detail_url"],
                    "excel_url": source["excel_url"],
                    "release_date": source.get("release_date", ""),
                    "sha256": sha256_file(local_path),
                    "local_path": str(local_path),
                }
            )

        source_signature = build_source_signature(downloaded)
        if old_meta.get("source_signature") == source_signature:
            print("No change: source signature unchanged.")
            return 0

        updated_at_utc = now_utc_iso()
        parts: list[pd.DataFrame] = []
        processed_sources: list[dict[str, str]] = []
        for item in downloaded:
            local_path = Path(item["local_path"])
            period_key = detect_period_key_from_sheet(local_path, item["period_key"])
            parsed = parse_airport_rows(
                path=local_path,
                period_key=period_key,
                source_url=item["excel_url"],
                updated_at_utc=updated_at_utc,
            )
            if parsed.empty:
                continue
            parts.append(parsed)
            processed_sources.append(
                {
                    "period_key": period_key,
                    "title": item["title"],
                    "detail_url": item["detail_url"],
                    "excel_url": item["excel_url"],
                    "sha256": item["sha256"],
                    "release_date": item.get("release_date", ""),
                }
            )

        if not parts:
            raise RuntimeError("No rows matched configured major airport mapping.")

        df = pd.concat(parts, ignore_index=True)
        df["arrivals"] = pd.to_numeric(df["arrivals"], errors="coerce").fillna(0).astype(int)
        df = (
            df.groupby(
                [
                    "period_key",
                    "airport_name_raw",
                    "airport_name",
                    "airport_code",
                    "unit",
                    "source_name",
                    "source_url",
                    "updated_at_utc",
                ],
                as_index=False,
            )["arrivals"]
            .sum()
            .copy()
        )
        df = df.sort_values(["period_key", "arrivals"], ascending=[True, False]).reset_index(
            drop=True
        )

        build_sqlite(df, SQLITE_PATH)

        latest = sorted(processed_sources, key=lambda x: x["period_key"], reverse=True)[0]
        period_values = sorted(df["period_key"].dropna().astype(str).unique().tolist())
        meta = {
            "source_name": SOURCE_NAME,
            "source_list_url": SOURCE_LIST_BASE_URL,
            "source_signature": source_signature,
            "source_excel_url": latest["excel_url"],
            "source_sha256": latest["sha256"],
            "fetched_at_utc": updated_at_utc,
            "period_min": period_values[0] if period_values else "",
            "period_max": period_values[-1] if period_values else "",
            "row_count": int(len(df)),
            "airports": (
                df[["airport_code", "airport_name"]]
                .drop_duplicates()
                .sort_values("airport_code")
                .to_dict(orient="records")
            ),
            "processed_sources": sorted(
                processed_sources, key=lambda x: x["period_key"], reverse=True
            ),
            "unit": UNIT,
        }
        save_meta(meta)

    print(
        f"Updated {TABLE_NAME}: periods={df['period_key'].nunique()} "
        f"airports={df['airport_code'].nunique()} rows={len(df)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
