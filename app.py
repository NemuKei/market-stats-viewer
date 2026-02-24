from __future__ import annotations

import hashlib
import io
import json
import re
import sqlite3
from pathlib import Path
from typing import cast

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout

REPO_ROOT = Path(__file__).resolve().parent
DATA_DIR = REPO_ROOT / "data"
SQLITE_PATH = DATA_DIR / "market_stats.sqlite"
META_PATH = DATA_DIR / "meta.json"
STAY_FACILITY_OCCUPANCY_TABLE_NAME = "stay_facility_occupancy"

STAY_VIEW_MODE_NIGHTS = "都道府県別 延べ宿泊者数"
STAY_VIEW_MODE_FACILITY_OCCUPANCY = "全国 宿泊施設種別 客室稼働率"
FACILITY_TYPE_DISPLAY_ORDER = [
    "計",
    "旅館",
    "リゾートホテル",
    "ビジネスホテル",
    "シティホテル",
    "簡易宿所",
    "会社・団体の宿泊所",
]

REGION_PREF_CODES = {
    "北海道": ["01"],
    "東北": ["02", "03", "04", "05", "06", "07"],
    "関東": ["08", "09", "10", "11", "12", "13", "14"],
    "中部": ["15", "16", "17", "18", "19", "20", "21", "22", "23"],
    "近畿（関西）": ["24", "25", "26", "27", "28", "29", "30"],
    "中国": ["31", "32", "33", "34", "35"],
    "四国": ["36", "37", "38", "39"],
    "九州・沖縄": ["40", "41", "42", "43", "44", "45", "46", "47"],
}
PREF_CODE_NAME_MAP = {
    "01": "北海道",
    "02": "青森県",
    "03": "岩手県",
    "04": "宮城県",
    "05": "秋田県",
    "06": "山形県",
    "07": "福島県",
    "08": "茨城県",
    "09": "栃木県",
    "10": "群馬県",
    "11": "埼玉県",
    "12": "千葉県",
    "13": "東京都",
    "14": "神奈川県",
    "15": "新潟県",
    "16": "富山県",
    "17": "石川県",
    "18": "福井県",
    "19": "山梨県",
    "20": "長野県",
    "21": "岐阜県",
    "22": "静岡県",
    "23": "愛知県",
    "24": "三重県",
    "25": "滋賀県",
    "26": "京都府",
    "27": "大阪府",
    "28": "兵庫県",
    "29": "奈良県",
    "30": "和歌山県",
    "31": "鳥取県",
    "32": "島根県",
    "33": "岡山県",
    "34": "広島県",
    "35": "山口県",
    "36": "徳島県",
    "37": "香川県",
    "38": "愛媛県",
    "39": "高知県",
    "40": "福岡県",
    "41": "佐賀県",
    "42": "長崎県",
    "43": "熊本県",
    "44": "大分県",
    "45": "宮崎県",
    "46": "鹿児島県",
    "47": "沖縄県",
}
REGION_PREF_NAMES = {
    region: [
        PREF_CODE_NAME_MAP.get(code, "")
        for code in pref_codes
        if PREF_CODE_NAME_MAP.get(code, "")
    ]
    for region, pref_codes in REGION_PREF_CODES.items()
}
TIME_SERIES_METRICS = {
    "国内+海外（積み上げ）": "stacked",
    "全体": "total",
    "国内": "jp",
    "海外": "foreign",
}
ANNUAL_METRICS = {"全体": "total", "国内": "jp", "海外": "foreign"}
CHART_VALUE_MODES = {
    "月次": "monthly",
    "年計推移（表記月起点・直近12か月ローリング）": "rolling12",
}


@st.cache_data(show_spinner=False)
def load_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query("SELECT * FROM market_stats", conn)
        df["pref_code"] = df["pref_code"].astype(str).str.zfill(2)
        df["ym"] = df["ym"].astype(str)
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_stay_facility_occupancy_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(
                f"SELECT * FROM {STAY_FACILITY_OCCUPANCY_TABLE_NAME}", conn
            )
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    df["ym"] = df["ym"].astype(str)
    if "pref_code" in df.columns:
        df["pref_code"] = df["pref_code"].astype(str).str.zfill(2)
    else:
        df["pref_code"] = "00"
    if "pref_name" in df.columns:
        df["pref_name"] = df["pref_name"].astype(str)
    else:
        df["pref_name"] = "全国"
    df["facility_type"] = (
        df["facility_type"]
        .astype(str)
        .str.replace(r"\s+", "", regex=True)
        .str.replace("\u3000", "", regex=False)
    )
    df["occupancy_rate"] = pd.to_numeric(df["occupancy_rate"], errors="coerce")
    df = df.dropna(subset=["occupancy_rate"]).copy()
    return df.sort_values(["ym", "facility_type"]).reset_index(drop=True)


def load_meta() -> dict:
    if not META_PATH.exists():
        return {}
    return json.loads(META_PATH.read_text(encoding="utf-8"))


def add_year_month_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["year"] = out["ym"].str.slice(0, 4).astype(int)
    out["month"] = out["ym"].str.slice(5, 7).astype(int)
    out["month_label"] = out["month"].map(lambda m: f"{m:02d}")
    return out


@st.cache_data(show_spinner=False)
def get_scope_dataframe(
    df: pd.DataFrame, scope_type: str, scope_id: str
) -> pd.DataFrame:
    if scope_type == "pref":
        out = df[df["pref_code"] == scope_id].copy()
        return out.sort_values("ym").reset_index(drop=True)

    pref_codes = REGION_PREF_CODES.get(scope_id, [])
    work = df[df["pref_code"].isin(pref_codes)].copy()
    if work.empty:
        return pd.DataFrame(columns=df.columns)

    grouped = (
        work.groupby("ym", as_index=False)[["total", "jp", "foreign"]]
        .sum()
        .sort_values("ym")
    )
    grouped["pref_code"] = scope_id
    grouped["pref_name"] = scope_id
    grouped = grouped[["ym", "pref_code", "pref_name", "foreign", "jp", "total"]]
    return grouped.reset_index(drop=True)


def ym_to_int(ym: str) -> int:
    return int(ym[:4]) * 100 + int(ym[5:7])


def build_ym(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def clamp_ym_to_available_range(ym: str, min_ym: str, max_ym: str) -> tuple[str, bool]:
    if ym < min_ym:
        return min_ym, True
    if ym > max_ym:
        return max_ym, True
    return ym, False


def sanitize_for_filename(value: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z_-]+", "_", value.strip()).strip("_")
    if safe:
        return safe
    suffix = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8]
    return f"scope_{suffix}"


def _ordered_pref_options(pref_options: list[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for _, prefs in REGION_PREF_NAMES.items():
        for pref in prefs:
            if pref in pref_options and pref not in seen:
                ordered.append(pref)
                seen.add(pref)
    for pref in pref_options:
        if pref not in seen:
            ordered.append(pref)
            seen.add(pref)
    return ordered


def render_pref_toggles_two_step(
    pref_options: list[str],
    pref_selection_key: str,
    pref_toggle_key_prefix: str,
    region_selection_key: str,
    region_toggle_key_prefix: str,
) -> list[str]:
    pref_options = _ordered_pref_options(pref_options)
    pref_set = set(pref_options)
    available_regions = [
        region
        for region, prefs in REGION_PREF_NAMES.items()
        if any(pref in pref_set for pref in prefs)
    ]

    st.markdown("**地方（複数選択）**")
    st.caption("未選択時は全地方（全都道府県）を表示します。")

    if region_selection_key not in st.session_state:
        st.session_state[region_selection_key] = []
    selected_region_seed = {
        str(v)
        for v in st.session_state.get(region_selection_key, [])
        if str(v) in available_regions
    }

    selected_regions: list[str] = []
    if available_regions:
        region_columns = st.columns(min(8, len(available_regions)))
        for idx, region in enumerate(available_regions):
            region_key = f"{region_toggle_key_prefix}_{region}"
            if region_key not in st.session_state:
                st.session_state[region_key] = region in selected_region_seed
            with region_columns[idx % len(region_columns)]:
                is_selected = st.toggle(region, key=region_key)
            if is_selected:
                selected_regions.append(region)
    st.session_state[region_selection_key] = selected_regions

    visible_prefs: list[str]
    if selected_regions:
        visible_prefs = []
        for region in available_regions:
            if region not in selected_regions:
                continue
            visible_prefs.extend(
                [pref for pref in REGION_PREF_NAMES[region] if pref in pref_set]
            )
    else:
        visible_prefs = pref_options

    st.markdown("**都道府県（複数選択）**")
    st.caption("未選択時は全都道府県を対象にします。")

    if pref_selection_key not in st.session_state:
        st.session_state[pref_selection_key] = []
    selected_pref_seed = {
        str(v)
        for v in st.session_state.get(pref_selection_key, [])
        if str(v) in visible_prefs
    }

    selected_prefs: list[str] = []
    if visible_prefs:
        pref_columns = st.columns(min(8, len(visible_prefs)))
        for idx, pref in enumerate(visible_prefs):
            pref_key = f"{pref_toggle_key_prefix}_{pref}"
            if pref_key not in st.session_state:
                st.session_state[pref_key] = pref in selected_pref_seed
            with pref_columns[idx % len(pref_columns)]:
                is_selected = st.toggle(pref, key=pref_key)
            if is_selected:
                selected_prefs.append(pref)

    st.session_state[pref_selection_key] = selected_prefs
    return selected_prefs


def normalize_selected_years(
    selected_years: list[int] | tuple[int, ...], available_years: list[int]
) -> list[int]:
    if not available_years:
        return []
    normalized = sorted({int(y) for y in selected_years if int(y) in available_years})
    if normalized:
        return normalized
    return available_years[-4:] if len(available_years) > 4 else available_years


def apply_rolling_12m(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    work = df.sort_values("ym").copy()
    target_cols = ["total", "jp", "foreign"]
    for col in target_cols:
        work[col] = work[col].rolling(window=12, min_periods=12).sum()
    work = work.dropna(subset=target_cols).copy()
    for col in target_cols:
        work[col] = work[col].round().astype("int64")
    return work.reset_index(drop=True)


def get_chart_source_dataframes(
    df_scope_all: pd.DataFrame, ym_from: str, ym_to: str, chart_value_mode_label: str
) -> tuple[pd.DataFrame, pd.DataFrame]:
    mode = CHART_VALUE_MODES.get(chart_value_mode_label, "monthly")
    if mode == "rolling12":
        chart_all = apply_rolling_12m(df_scope_all)
    else:
        chart_all = df_scope_all.sort_values("ym").reset_index(drop=True)

    chart_filtered = chart_all[
        (chart_all["ym"] >= ym_from) & (chart_all["ym"] <= ym_to)
    ]
    return chart_filtered.reset_index(drop=True), chart_all


def write_helper_table(
    ws, df: pd.DataFrame, start_row: int, start_col: int = 1
) -> tuple[int, int]:
    headers = df.columns.tolist()
    for j, header in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=j, value=header)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=value)
    end_row = start_row if df.empty else start_row + len(df)
    end_col = start_col + len(headers) - 1
    return end_row, end_col


def try_set_attr(obj, attr_name: str, value) -> None:
    try:
        setattr(obj, attr_name, value)
    except Exception:
        pass


def assign_axis_ids(chart, x_id: int, y_id: int) -> None:
    chart.axId = [x_id, y_id]
    chart.x_axis.axId = x_id
    chart.y_axis.axId = y_id
    chart.x_axis.crossAx = y_id
    chart.y_axis.crossAx = x_id


def build_excel_report_bytes(
    df_table: pd.DataFrame,
    df_chart_filtered: pd.DataFrame,
    df_chart_all: pd.DataFrame,
    selection_state: dict,
) -> bytes:
    time_series_label = selection_state["time_series_label"]
    annual_metric_label = selection_state["annual_metric_label"]
    annual_years = selection_state["annual_years"]
    chart_value_mode_label = selection_state["chart_value_mode_label"]
    is_rolling = CHART_VALUE_MODES.get(chart_value_mode_label) == "rolling12"
    rolling_suffix = "・12か月ローリング" if is_rolling else ""

    data_sheet_df = df_table[
        ["ym", "pref_code", "pref_name", "total", "jp", "foreign"]
    ].copy()
    data_sheet_df = data_sheet_df.rename(
        columns={
            "ym": "年月",
            "pref_code": "地域コード",
            "pref_name": "地域名",
            "total": "全体",
            "jp": "国内",
            "foreign": "海外",
        }
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        data_sheet_df.to_excel(writer, sheet_name="data", index=False)
        workbook = writer.book
        charts_ws = workbook.create_sheet("charts")

        charts_ws["A1"] = "Exported charts"
        helper_col = 30  # AD列付近に補助表を置き、見た目への干渉を避ける

        # Time-series helper and chart
        ts_base = add_year_month_columns(
            df_chart_filtered[["ym", "total", "jp", "foreign"]]
        ).sort_values("ym")
        ts_mode = TIME_SERIES_METRICS.get(time_series_label, "stacked")
        if ts_mode == "stacked":
            ts_helper = ts_base[["ym", "jp", "foreign"]].rename(
                columns={"ym": "年月", "jp": "国内", "foreign": "海外"}
            )
            ts_title = f"時系列（積み上げ{rolling_suffix}）"
            ts_grouping = "stacked"
        else:
            ts_helper = ts_base[["ym", ts_mode]].rename(
                columns={"ym": "年月", ts_mode: time_series_label}
            )
            ts_title = f"時系列（{time_series_label}{rolling_suffix}）"
            ts_grouping = "clustered"

        ts_start_row = 2
        ts_end_row, ts_end_col = write_helper_table(
            charts_ws, ts_helper, ts_start_row, start_col=helper_col
        )
        if not ts_helper.empty:
            ts_chart = BarChart()
            ts_chart.type = "col"
            ts_chart.grouping = ts_grouping
            assign_axis_ids(ts_chart, 10, 100)
            if ts_grouping == "stacked":
                ts_chart.overlap = 100
            ts_chart.title = ts_title
            ts_chart.y_axis.title = "延べ宿泊者数"
            ts_chart.x_axis.title = None
            ts_chart.legend.position = "r"
            ts_chart.legend.overlay = False
            ts_chart.layout = Layout(
                manualLayout=ManualLayout(x=0.04, y=0.08, w=0.78, h=0.78)
            )
            ts_data = Reference(
                charts_ws,
                min_col=helper_col + 1,
                max_col=ts_end_col,
                min_row=ts_start_row,
                max_row=ts_end_row,
            )
            ts_categories = Reference(
                charts_ws,
                min_col=helper_col,
                min_row=ts_start_row + 1,
                max_row=ts_end_row,
            )
            ts_chart.add_data(ts_data, titles_from_data=True)
            ts_chart.set_categories(ts_categories)
            ts_chart.x_axis.delete = False
            ts_chart.y_axis.delete = False
            ts_chart.x_axis.tickLblPos = "low"
            ts_chart.x_axis.tickLblSkip = 3
            try_set_attr(ts_chart.x_axis, "tickMarkSkip", 3)
            ts_chart.x_axis.majorTickMark = "out"
            ts_chart.y_axis.majorTickMark = "out"
            try_set_attr(ts_chart.y_axis, "numFmt", "#,##0")
            ts_chart.y_axis.majorGridlines = ChartLines()
            ts_chart.width = 26
            ts_chart.height = 11
            charts_ws.add_chart(ts_chart, "A2")
        else:
            charts_ws["A2"] = "時系列グラフ: データなし"

        # Annual comparison helper and chart
        annual_base = add_year_month_columns(
            df_chart_all[["ym", "total", "jp", "foreign"]]
        )
        available_years = sorted(annual_base["year"].unique().tolist())
        years_for_chart = normalize_selected_years(annual_years, available_years)
        annual_col = ANNUAL_METRICS.get(annual_metric_label, "total")

        annual_pivot = (
            annual_base[annual_base["year"].isin(years_for_chart)]
            .pivot_table(
                index="month", columns="year", values=annual_col, aggfunc="sum"
            )
            .reindex(range(1, 13))
        )
        for y in years_for_chart:
            if y not in annual_pivot.columns:
                annual_pivot[y] = float("nan")
        annual_pivot = (
            annual_pivot[years_for_chart] if years_for_chart else annual_pivot
        )
        if not is_rolling:
            annual_pivot = annual_pivot.fillna(0)
        annual_pivot.index = [f"{m:02d}" for m in annual_pivot.index]
        annual_helper = annual_pivot.reset_index().rename(
            columns={"index": "月", "month": "月"}
        )
        annual_helper.columns = ["月"] + [str(y) for y in years_for_chart]

        annual_start_row = max(30, ts_end_row + 3)
        annual_end_row, annual_end_col = write_helper_table(
            charts_ws, annual_helper, annual_start_row, start_col=helper_col
        )

        # Guard: overlap regression detector for time-series category column.
        ts_category_values = [
            charts_ws.cell(row=r, column=helper_col).value
            for r in range(ts_start_row + 1, ts_end_row + 1)
        ]
        if any(v == "月" for v in ts_category_values):
            raise RuntimeError(
                "Helper table overlap detected: time-series categories contain annual header."
            )

        if years_for_chart:
            annual_chart = BarChart()
            annual_chart.type = "col"
            annual_chart.grouping = "clustered"
            assign_axis_ids(annual_chart, 20, 200)
            annual_chart.title = (
                f"年別同月比較（{annual_metric_label}{rolling_suffix}）"
            )
            annual_chart.y_axis.title = "延べ宿泊者数"
            annual_chart.x_axis.title = None
            annual_chart.legend.position = "r"
            annual_chart.legend.overlay = False
            annual_chart.layout = Layout(
                manualLayout=ManualLayout(x=0.04, y=0.08, w=0.78, h=0.78)
            )
            annual_data = Reference(
                charts_ws,
                min_col=helper_col + 1,
                max_col=annual_end_col,
                min_row=annual_start_row,
                max_row=annual_end_row,
            )
            annual_categories = Reference(
                charts_ws,
                min_col=helper_col,
                min_row=annual_start_row + 1,
                max_row=annual_end_row,
            )
            annual_chart.add_data(annual_data, titles_from_data=True)
            annual_chart.set_categories(annual_categories)
            annual_chart.x_axis.delete = False
            annual_chart.y_axis.delete = False
            annual_chart.x_axis.tickLblPos = "low"
            annual_chart.x_axis.tickLblSkip = 1
            try_set_attr(annual_chart.x_axis, "tickMarkSkip", 1)
            annual_chart.x_axis.majorTickMark = "out"
            annual_chart.y_axis.majorTickMark = "out"
            try_set_attr(annual_chart.y_axis, "numFmt", "#,##0")
            annual_chart.y_axis.majorGridlines = ChartLines()
            annual_chart.width = 26
            annual_chart.height = 12
            charts_ws.add_chart(annual_chart, "A30")
        else:
            charts_ws["A30"] = "年別同月比較グラフ: データなし"

    buffer.seek(0)
    return buffer.getvalue()


def build_time_series_chart(
    df_filtered: pd.DataFrame, metric_mode: str
) -> alt.Chart | alt.LayerChart:
    work = add_year_month_columns(df_filtered[["ym", "total", "jp", "foreign"]])
    ym_sort = sorted(work["ym"].unique().tolist())

    if TIME_SERIES_METRICS[metric_mode] == "stacked":
        work["total"] = work["jp"] + work["foreign"]
        long_df = work.melt(
            id_vars=["ym", "year", "month", "total"],
            value_vars=["jp", "foreign"],
            var_name="metric_key",
            value_name="value",
        )
        metric_labels = {"jp": "国内", "foreign": "海外"}
        long_df["metric"] = long_df["metric_key"].map(metric_labels)
        long_df["stack_order"] = long_df["metric_key"].map({"foreign": 0, "jp": 1})
        long_df["share_label"] = long_df.apply(
            lambda r: (
                f"{int(round((r['value'] / r['total']) * 100))}%"
                if r["total"] > 0
                else ""
            ),
            axis=1,
        )

        label_base = work[["ym", "year", "month", "total", "jp", "foreign"]].copy()
        jp_label_df = label_base.copy()
        jp_label_df["metric"] = metric_labels["jp"]
        jp_label_df["y_center"] = jp_label_df["foreign"] + (jp_label_df["jp"] / 2)
        jp_label_df["share_label"] = jp_label_df.apply(
            lambda r: (
                f"{int(round((r['jp'] / r['total']) * 100))}%" if r["total"] > 0 else ""
            ),
            axis=1,
        )
        foreign_label_df = label_base.copy()
        foreign_label_df["metric"] = metric_labels["foreign"]
        foreign_label_df["y_center"] = foreign_label_df["foreign"] / 2
        foreign_label_df["share_label"] = foreign_label_df.apply(
            lambda r: (
                f"{int(round((r['foreign'] / r['total']) * 100))}%"
                if r["total"] > 0
                else ""
            ),
            axis=1,
        )
        share_df = pd.concat([jp_label_df, foreign_label_df], ignore_index=True)

        bars = (
            alt.Chart(long_df)
            .mark_bar()
            .encode(
                x=alt.X("ym:N", title="年月", sort=ym_sort),
                y=alt.Y("value:Q", title="延べ宿泊者数"),
                color=alt.Color(
                    "metric:N",
                    title="区分",
                    sort=["国内", "海外"],
                    scale=alt.Scale(domain=["国内", "海外"]),
                ),
                order=alt.Order("stack_order:Q", sort="ascending"),
                tooltip=[
                    alt.Tooltip("ym:N", title="年月"),
                    alt.Tooltip("year:Q", title="年"),
                    alt.Tooltip("month:Q", title="月"),
                    alt.Tooltip("metric:N", title="区分"),
                    alt.Tooltip("value:Q", title="値", format=",.0f"),
                    alt.Tooltip("share_label:N", title="シェア"),
                ],
            )
        )

        share_text = (
            alt.Chart(share_df)
            .mark_text(color="white", fontSize=10)
            .encode(
                x=alt.X("ym:N", sort=ym_sort),
                y=alt.Y("y_center:Q"),
                text=alt.Text("share_label:N"),
                detail=alt.Detail("metric:N"),
            )
        )

        return bars + share_text

    metric_col = TIME_SERIES_METRICS[metric_mode]
    single_df = work[["ym", "year", "month", metric_col]].copy()
    single_df = single_df.rename(columns={metric_col: "value"})
    single_df["metric"] = metric_mode

    return (
        alt.Chart(single_df)
        .mark_bar()
        .encode(
            x=alt.X("ym:N", title="年月", sort=ym_sort),
            y=alt.Y("value:Q", title="延べ宿泊者数"),
            color=alt.value("#4C78A8"),
            tooltip=[
                alt.Tooltip("ym:N", title="年月"),
                alt.Tooltip("year:Q", title="年"),
                alt.Tooltip("month:Q", title="月"),
                alt.Tooltip("metric:N", title="区分"),
                alt.Tooltip("value:Q", title="値", format=",.0f"),
            ],
        )
    )


def build_yearly_month_compare_chart(
    df_scope_all: pd.DataFrame, metric_col: str, selected_years: list[int]
) -> alt.Chart:
    month_sort = [f"{m:02d}" for m in range(1, 13)]
    work = add_year_month_columns(df_scope_all[["ym", "total", "jp", "foreign"]])
    work = work[work["year"].isin(selected_years)].copy()

    return (
        alt.Chart(work)
        .mark_bar()
        .encode(
            x=alt.X("month_label:N", title="月", sort=month_sort),
            xOffset=alt.XOffset("year:N", sort=selected_years),
            y=alt.Y(f"{metric_col}:Q", title="延べ宿泊者数"),
            color=alt.Color("year:N", title="年", sort=selected_years),
            tooltip=[
                alt.Tooltip("ym:N", title="年月"),
                alt.Tooltip("year:Q", title="年"),
                alt.Tooltip("month:Q", title="月"),
                alt.Tooltip(f"{metric_col}:Q", title="値", format=",.0f"),
            ],
        )
    )


def get_ordered_facility_types(facility_types: list[str]) -> list[str]:
    unique = sorted({str(v).strip() for v in facility_types if str(v).strip()})
    order_map = {name: idx for idx, name in enumerate(FACILITY_TYPE_DISPLAY_ORDER)}
    return sorted(unique, key=lambda x: (order_map.get(x, 999), x))


def build_facility_occupancy_timeseries_chart(df_filtered: pd.DataFrame) -> alt.Chart:
    ym_sort = sorted(df_filtered["ym"].astype(str).unique().tolist())
    return (
        alt.Chart(df_filtered)
        .mark_line(point=True)
        .encode(
            x=alt.X("ym:N", title="年月", sort=ym_sort),
            y=alt.Y(
                "occupancy_rate:Q",
                title="客室稼働率（%）",
                scale=alt.Scale(domain=[0, 100]),
            ),
            color=alt.Color("facility_type:N", title="宿泊施設種別"),
            tooltip=[
                alt.Tooltip("ym:N", title="年月"),
                alt.Tooltip("facility_type:N", title="宿泊施設種別"),
                alt.Tooltip("occupancy_rate:Q", title="稼働率（%）", format=".1f"),
            ],
        )
    )


def build_facility_occupancy_fiscal_compare_chart(
    df: pd.DataFrame, fiscal_years: list[int]
) -> alt.Chart:
    if df.empty:
        return alt.Chart(pd.DataFrame(columns=["fiscal_month_label", "occupancy_rate"]))

    work = add_year_month_columns(df[["ym", "occupancy_rate"]])
    work["fiscal_year"] = work.apply(
        lambda r: int(r["year"]) if int(r["month"]) >= 4 else int(r["year"]) - 1,
        axis=1,
    )
    work["fiscal_month"] = work["month"].astype(int)
    month_order = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    month_order_labels = [f"{m:02d}" for m in month_order]
    work["fiscal_month_label"] = work["fiscal_month"].map(lambda m: f"{m:02d}")
    work = work[work["fiscal_year"].isin(fiscal_years)].copy()

    grouped = (
        work.groupby(["fiscal_year", "fiscal_month_label"], as_index=False)[
            "occupancy_rate"
        ]
        .mean()
        .copy()
    )

    return (
        alt.Chart(grouped)
        .mark_bar()
        .encode(
            x=alt.X(
                "fiscal_month_label:N", title="月（年度）", sort=month_order_labels
            ),
            xOffset=alt.XOffset("fiscal_year:N", sort=fiscal_years),
            y=alt.Y(
                "occupancy_rate:Q",
                title="客室稼働率（%）",
                scale=alt.Scale(domain=[0, 100]),
            ),
            color=alt.Color("fiscal_year:N", title="年度", sort=fiscal_years),
            tooltip=[
                alt.Tooltip("fiscal_year:N", title="年度"),
                alt.Tooltip("fiscal_month_label:N", title="月"),
                alt.Tooltip("occupancy_rate:Q", title="稼働率（%）", format=".1f"),
            ],
        )
    )


def build_facility_occupancy_fiscal_type_compare_chart(
    df: pd.DataFrame, fiscal_year: int, facility_types: list[str]
) -> alt.Chart:
    if df.empty or not facility_types:
        return alt.Chart(
            pd.DataFrame(
                columns=["fiscal_month_label", "facility_type", "occupancy_rate"]
            )
        )

    work = add_year_month_columns(df[["ym", "facility_type", "occupancy_rate"]])
    work["fiscal_year"] = work.apply(
        lambda r: int(r["year"]) if int(r["month"]) >= 4 else int(r["year"]) - 1,
        axis=1,
    )
    work["fiscal_month"] = work["month"].astype(int)
    month_order = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    month_order_labels = [f"{m:02d}" for m in month_order]
    work["fiscal_month_label"] = work["fiscal_month"].map(lambda m: f"{m:02d}")
    work = work[
        (work["fiscal_year"] == int(fiscal_year))
        & (work["facility_type"].isin(facility_types))
    ].copy()

    grouped = (
        work.groupby(["facility_type", "fiscal_month_label"], as_index=False)[
            "occupancy_rate"
        ]
        .mean()
        .copy()
    )
    grouped["fiscal_year"] = int(fiscal_year)

    return (
        alt.Chart(grouped)
        .mark_bar()
        .encode(
            x=alt.X(
                "fiscal_month_label:N", title="月（年度）", sort=month_order_labels
            ),
            xOffset=alt.XOffset("facility_type:N", sort=facility_types),
            y=alt.Y(
                "occupancy_rate:Q",
                title="客室稼働率（%）",
                scale=alt.Scale(domain=[0, 100]),
            ),
            color=alt.Color(
                "facility_type:N", title="宿泊施設種別", sort=facility_types
            ),
            tooltip=[
                alt.Tooltip("fiscal_year:N", title="年度"),
                alt.Tooltip("facility_type:N", title="宿泊施設種別"),
                alt.Tooltip("fiscal_month_label:N", title="月"),
                alt.Tooltip("occupancy_rate:Q", title="稼働率（%）", format=".1f"),
            ],
        )
    )


def render_stay_facility_occupancy_view(meta: dict) -> None:
    st.subheader("宿泊施設種別 客室稼働率")
    if meta and meta.get("facility_occupancy_rows"):
        st.caption(
            f"客室稼働率データ範囲: {meta.get('facility_occupancy_min_ym')}～"
            f"{meta.get('facility_occupancy_max_ym')} / rows: {meta.get('facility_occupancy_rows')}"
        )

    df = load_stay_facility_occupancy_data()
    if df.empty:
        st.error(
            "客室稼働率データがありません。先に python -m scripts.update_data を実行してください。"
        )
        return

    top_col1, top_col2 = st.columns([2, 3])
    with top_col1:
        scope_label = st.radio(
            "地域区分",
            ["全国", "都道府県"],
            horizontal=True,
            key="facility_occ_scope_label",
        )
    scope_df = df[df["pref_code"] == "00"].copy()
    scope_name = "全国"
    if scope_label == "都道府県":
        prefs = (
            df[df["pref_code"] != "00"][["pref_code", "pref_name"]]
            .drop_duplicates()
            .sort_values("pref_code")
        )
        pref_labels = prefs.apply(
            lambda r: f"{r['pref_code']} {r['pref_name']}", axis=1
        ).tolist()
        pref_map = dict(zip(pref_labels, prefs["pref_code"].tolist()))
        with top_col2:
            pref_sel = st.selectbox(
                "都道府県",
                pref_labels,
                index=0,
                key="facility_occ_pref",
            )
        pref_code = pref_map[pref_sel]
        scope_df = df[df["pref_code"] == pref_code].copy()
        scope_name = pref_sel
    else:
        with top_col2:
            st.caption("対象: 全国")

    if scope_df.empty:
        st.info("選択した地域に客室稼働率データがありません。")
        return

    facility_options = get_ordered_facility_types(
        scope_df["facility_type"].astype(str).tolist()
    )
    if not facility_options:
        st.info("宿泊施設種別データがありません。")
        return

    filter_col1, filter_col2, filter_col3 = st.columns([3, 2, 2])
    with filter_col1:
        default_timeseries_types = facility_options
        selected_facility_types = st.multiselect(
            "宿泊施設種別",
            options=facility_options,
            default=default_timeseries_types,
            key="facility_occ_types_multi",
        )

    if not selected_facility_types:
        st.info("宿泊施設種別を1つ以上選択してください。")
        return

    target_df_all = scope_df[
        scope_df["facility_type"].isin(selected_facility_types)
    ].copy()
    ym_options = sorted(target_df_all["ym"].astype(str).unique().tolist())
    if not ym_options:
        st.info("選択した条件のデータがありません。")
        return

    min_ym = ym_options[0]
    max_ym = ym_options[-1]
    default_ym_from = ym_options[max(0, len(ym_options) - 36)]
    default_ym_to = ym_options[-1]
    year_options = sorted(
        target_df_all["ym"].str.slice(0, 4).astype(int).unique().tolist()
    )
    month_options = list(range(1, 13))

    def _fmt_month(v: int) -> str:
        return f"{v:02d}"

    default_from_year = int(default_ym_from[:4])
    default_from_month = int(default_ym_from[5:7])
    default_to_year = int(default_ym_to[:4])
    default_to_month = int(default_ym_to[5:7])

    with filter_col2:
        sy_col, sm_col = st.columns(2)
        with sy_col:
            start_year = st.selectbox(
                "開始（年）",
                year_options,
                index=year_options.index(default_from_year),
                key="facility_occ_start_year",
            )
        with sm_col:
            start_month = st.selectbox(
                "開始（月）",
                month_options,
                index=month_options.index(default_from_month),
                format_func=_fmt_month,
                key="facility_occ_start_month",
            )

    with filter_col3:
        ey_col, em_col = st.columns(2)
        with ey_col:
            end_year = st.selectbox(
                "終了（年）",
                year_options,
                index=year_options.index(default_to_year),
                key="facility_occ_end_year",
            )
        with em_col:
            end_month = st.selectbox(
                "終了（月）",
                month_options,
                index=month_options.index(default_to_month),
                format_func=_fmt_month,
                key="facility_occ_end_month",
            )

    ym_from = build_ym(start_year, start_month)
    ym_to = build_ym(end_year, end_month)
    ym_from, from_clamped = clamp_ym_to_available_range(ym_from, min_ym, max_ym)
    ym_to, to_clamped = clamp_ym_to_available_range(ym_to, min_ym, max_ym)
    if from_clamped:
        st.warning(f"開始年月をデータ範囲に合わせて {ym_from} に補正しました。")
    if to_clamped:
        st.warning(f"終了年月をデータ範囲に合わせて {ym_to} に補正しました。")
    if ym_to_int(ym_from) > ym_to_int(ym_to):
        ym_from, ym_to = ym_to, ym_from
        st.warning(
            f"開始年月と終了年月が逆だったため、{ym_from} ～ {ym_to} に入れ替えました。"
        )

    ranged_df = target_df_all[
        (target_df_all["ym"] >= ym_from) & (target_df_all["ym"] <= ym_to)
    ].copy()
    if ranged_df.empty:
        st.info("指定した期間にデータがありません。")
        return

    latest_ym = ranged_df["ym"].max()
    latest_series = cast(
        pd.Series,
        ranged_df.loc[ranged_df["ym"] == latest_ym, "occupancy_rate"],
    )
    latest_values = pd.to_numeric(latest_series, errors="coerce").dropna()
    latest_mean = float(latest_values.mean()) if not latest_values.empty else 0.0
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("対象", scope_name)
    metric_col2.metric("表示期間", f"{ym_from} ～ {ym_to}")
    metric_col3.metric(
        f"最新月平均（{latest_ym}）",
        f"{latest_mean:.1f}%",
    )

    st.subheader("時系列")
    line_chart = build_facility_occupancy_timeseries_chart(ranged_df).properties(
        height=380
    )
    st.altair_chart(line_chart, use_container_width=True)

    st.subheader("年度比較（4月～翌3月）")
    fiscal_compare_mode = st.radio(
        "比較軸",
        ["年度比較（種別固定）", "種別比較（年度固定）"],
        horizontal=True,
        key="facility_occ_fiscal_compare_mode",
    )

    fiscal_all_scope = add_year_month_columns(scope_df[["ym"]].drop_duplicates().copy())
    fiscal_all_scope["fiscal_year"] = fiscal_all_scope.apply(
        lambda r: int(r["year"]) if int(r["month"]) >= 4 else int(r["year"]) - 1,
        axis=1,
    )
    fiscal_year_options_scope = sorted(
        fiscal_all_scope["fiscal_year"].astype(int).unique().tolist()
    )

    if fiscal_compare_mode == "年度比較（種別固定）":
        fiscal_default_type = "計" if "計" in facility_options else facility_options[0]
        fiscal_filter_col1, fiscal_filter_col2 = st.columns([3, 4])
        with fiscal_filter_col1:
            fiscal_facility_type = st.selectbox(
                "宿泊施設種別（年度比較）",
                options=facility_options,
                index=facility_options.index(fiscal_default_type),
                key="facility_occ_fiscal_type",
            )

        fiscal_target_df_all = scope_df[
            scope_df["facility_type"] == fiscal_facility_type
        ].copy()
        fiscal_all = add_year_month_columns(
            fiscal_target_df_all[["ym", "occupancy_rate"]]
        )
        fiscal_all["fiscal_year"] = fiscal_all.apply(
            lambda r: int(r["year"]) if int(r["month"]) >= 4 else int(r["year"]) - 1,
            axis=1,
        )
        fiscal_year_options = sorted(
            fiscal_all["fiscal_year"].astype(int).unique().tolist()
        )
        default_fiscal_years = (
            fiscal_year_options[-4:]
            if len(fiscal_year_options) > 4
            else fiscal_year_options
        )

        with fiscal_filter_col2:
            selected_fiscal_years = st.multiselect(
                "年度",
                options=fiscal_year_options,
                default=default_fiscal_years,
                key="facility_occ_fiscal_years",
            )
        if not selected_fiscal_years:
            st.info("年度を1つ以上選択してください。")
        else:
            fiscal_chart = build_facility_occupancy_fiscal_compare_chart(
                fiscal_target_df_all, selected_fiscal_years
            ).properties(height=380)
            st.altair_chart(fiscal_chart, use_container_width=True)
    else:
        fiscal_filter_col1, fiscal_filter_col2 = st.columns([3, 4])
        with fiscal_filter_col1:
            selected_fiscal_year = st.selectbox(
                "年度（種別比較）",
                options=fiscal_year_options_scope,
                index=len(fiscal_year_options_scope) - 1,
                key="facility_occ_fiscal_year_single",
            )
        with fiscal_filter_col2:
            selected_compare_types = st.multiselect(
                "宿泊施設種別（種別比較）",
                options=facility_options,
                default=facility_options,
                key="facility_occ_fiscal_types_multi",
            )

        if not selected_compare_types:
            st.info("宿泊施設種別を1つ以上選択してください。")
        else:
            fiscal_chart = build_facility_occupancy_fiscal_type_compare_chart(
                scope_df, int(selected_fiscal_year), selected_compare_types
            ).properties(height=380)
            st.altair_chart(fiscal_chart, use_container_width=True)

    st.subheader("表")
    table_df = ranged_df[
        ["ym", "pref_code", "pref_name", "facility_type", "occupancy_rate"]
    ].copy()
    table_df = table_df.sort_values(["ym", "facility_type"]).reset_index(drop=True)
    table_df["occupancy_rate"] = table_df["occupancy_rate"].round(1)
    table_df = table_df.rename(
        columns={
            "ym": "年月",
            "pref_code": "都道府県コード",
            "pref_name": "都道府県",
            "facility_type": "宿泊施設種別",
            "occupancy_rate": "客室稼働率（%）",
        }
    )
    st.dataframe(table_df, use_container_width=True, hide_index=True, height=520)


def render_stay_stats_view() -> None:
    st.title("宿泊旅行統計調査")

    meta = load_meta()
    if meta:
        st.caption(
            f"最終取得（UTC）: {meta.get('fetched_at_utc')} / "
            f"範囲: {meta.get('min_ym')}〜{meta.get('max_ym')} / "
            f"rows: {meta.get('rows')}"
        )

    view_mode = st.radio(
        "表示軸",
        [STAY_VIEW_MODE_NIGHTS, STAY_VIEW_MODE_FACILITY_OCCUPANCY],
        horizontal=True,
        key="stay_view_mode",
    )
    if view_mode == STAY_VIEW_MODE_FACILITY_OCCUPANCY:
        render_stay_facility_occupancy_view(meta)
        st.divider()
        st.caption("出典：観光庁『宿泊旅行統計調査』（4-2 客室稼働率（月別）を整形）")
        st.caption(
            "データは毎週自動更新です。取得元サイトの構造変更等により更新が遅れる場合があります。"
        )
        return

    st.subheader("延べ宿泊者数（全体 / 国内 / 海外）")

    df = load_data()
    if df.empty:
        st.error(
            "データがありません。先に python -m scripts.update_data を実行して data/ を生成してください。"
        )
        return

    # フィルタ
    col1, col2, col3, col4 = st.columns([2, 3, 4, 3])
    with col1:
        scope_label = st.radio("地域区分", ["都道府県", "地方"], horizontal=True)
    with col2:
        if scope_label == "都道府県":
            prefs = (
                df[["pref_code", "pref_name"]]
                .drop_duplicates()
                .sort_values("pref_code")
            )
            pref_label = prefs.apply(
                lambda r: f"{r['pref_code']} {r['pref_name']}", axis=1
            ).tolist()
            pref_map = dict(zip(pref_label, prefs["pref_code"].tolist()))
            pref_sel = st.selectbox("地域（全国/都道府県）", pref_label, index=0)
            scope_type = "pref"
            scope_id = pref_map[pref_sel]
        else:
            region_name = st.selectbox("地方", list(REGION_PREF_CODES.keys()), index=0)
            st.caption(
                f"対象都道府県コード: {', '.join(REGION_PREF_CODES[region_name])}"
            )
            scope_type = "region"
            scope_id = region_name

    d_scope_all = get_scope_dataframe(df, scope_type, scope_id)
    if d_scope_all.empty:
        st.error("選択した地域区分/地域に対応するデータがありません。")
        return

    ym_list = sorted(d_scope_all["ym"].unique().tolist())
    min_ym = ym_list[0]
    max_ym = ym_list[-1]
    default_ym_from = ym_list[max(0, len(ym_list) - 36)]
    default_ym_to = ym_list[-1]

    default_from_year = int(default_ym_from[:4])
    default_from_month = int(default_ym_from[5:7])
    default_to_year = int(default_ym_to[:4])
    default_to_month = int(default_ym_to[5:7])

    year_options = sorted(
        d_scope_all["ym"].str.slice(0, 4).astype(int).unique().tolist()
    )
    month_options = list(range(1, 13))

    def format_month(m: int) -> str:
        return f"{m:02d}"

    with col3:
        start_col, end_col = st.columns(2)
        with start_col:
            st.caption("開始")
            start_year = st.selectbox(
                "開始（年）",
                year_options,
                index=year_options.index(default_from_year),
            )
            start_month = st.selectbox(
                "開始（月）",
                month_options,
                index=month_options.index(default_from_month),
                format_func=format_month,
            )
        with end_col:
            st.caption("終了")
            end_year = st.selectbox(
                "終了（年）",
                year_options,
                index=year_options.index(default_to_year),
            )
            end_month = st.selectbox(
                "終了（月）",
                month_options,
                index=month_options.index(default_to_month),
                format_func=format_month,
            )

    with col4:
        show_mode = st.radio(
            "表示", ["表＋グラフ", "表のみ", "グラフのみ"], horizontal=True
        )

    ym_from = build_ym(start_year, start_month)
    ym_to = build_ym(end_year, end_month)

    ym_from, from_clamped = clamp_ym_to_available_range(ym_from, min_ym, max_ym)
    ym_to, to_clamped = clamp_ym_to_available_range(ym_to, min_ym, max_ym)
    if from_clamped:
        st.warning(f"開始年月をデータ範囲に合わせて {ym_from} に補正しました。")
    if to_clamped:
        st.warning(f"終了年月をデータ範囲に合わせて {ym_to} に補正しました。")

    if ym_to_int(ym_from) > ym_to_int(ym_to):
        ym_from, ym_to = ym_to, ym_from
        st.warning(
            f"開始年月と終了年月が逆だったため、{ym_from} ～ {ym_to} に入れ替えました。"
        )

    d = d_scope_all[
        (d_scope_all["ym"] >= ym_from) & (d_scope_all["ym"] <= ym_to)
    ].copy()
    d = d.sort_values("ym")

    # 表（年月縦）
    table = d[["ym", "total", "jp", "foreign"]].copy()
    table = table.rename(
        columns={"ym": "年月", "total": "全体", "jp": "国内", "foreign": "海外"}
    )
    scope_file_id = sanitize_for_filename(f"{scope_type}_{scope_id}")
    export_file_stem = f"market_stats_{scope_file_id}_{ym_from}_{ym_to}"
    chart_mode_options = [
        "時系列（積み上げ縦棒：国内＋海外）",
        "年別（同月比較：全体/国内/海外）",
    ]
    chart_value_mode_options = list(CHART_VALUE_MODES.keys())
    chart_mode = st.session_state.get("chart_mode_export", chart_mode_options[0])
    if chart_mode not in chart_mode_options:
        chart_mode = chart_mode_options[0]
    chart_value_mode_label = st.session_state.get(
        "chart_value_mode_export", chart_value_mode_options[0]
    )
    if chart_value_mode_label not in CHART_VALUE_MODES:
        chart_value_mode_label = chart_value_mode_options[0]
    ts_metric_label = st.session_state.get("ts_metric_export", "国内+海外（積み上げ）")
    if ts_metric_label not in TIME_SERIES_METRICS:
        ts_metric_label = "国内+海外（積み上げ）"
    annual_metric_label = st.session_state.get("annual_metric_export", "全体")
    if annual_metric_label not in ANNUAL_METRICS:
        annual_metric_label = "全体"
    chart_filtered, chart_scope_all = get_chart_source_dataframes(
        d_scope_all, ym_from, ym_to, chart_value_mode_label
    )
    chart_year_options = (
        sorted(chart_scope_all["ym"].str.slice(0, 4).astype(int).unique().tolist())
        if not chart_scope_all.empty
        else []
    )
    default_chart_years = (
        chart_year_options[-4:] if len(chart_year_options) > 4 else chart_year_options
    )
    selected_years_for_export = normalize_selected_years(
        st.session_state.get("annual_years_export", default_chart_years),
        chart_year_options,
    )

    chart_height = 520
    if show_mode in ["表＋グラフ", "グラフのみ"]:
        st.subheader("グラフ")
        chart_mode = st.radio(
            "チャートモード",
            chart_mode_options,
            key="chart_mode_export",
        )
        chart_value_mode_label = st.radio(
            "値の種類",
            chart_value_mode_options,
            horizontal=True,
            key="chart_value_mode_export",
        )
        chart_filtered, chart_scope_all = get_chart_source_dataframes(
            d_scope_all, ym_from, ym_to, chart_value_mode_label
        )
        chart_year_options = (
            sorted(chart_scope_all["ym"].str.slice(0, 4).astype(int).unique().tolist())
            if not chart_scope_all.empty
            else []
        )
        default_chart_years = (
            chart_year_options[-4:]
            if len(chart_year_options) > 4
            else chart_year_options
        )
        selected_years_for_export = normalize_selected_years(
            st.session_state.get("annual_years_export", default_chart_years),
            chart_year_options,
        )

        if chart_mode == "時系列（積み上げ縦棒：国内＋海外）":
            ts_metric_label = st.radio(
                "時系列の表示内容",
                list(TIME_SERIES_METRICS.keys()),
                horizontal=True,
                key="ts_metric_export",
            )
            if chart_filtered.empty:
                if CHART_VALUE_MODES.get(chart_value_mode_label) == "rolling12":
                    st.info(
                        "指定した期間にデータがありません。年計推移では各月で直近12か月分が必要です。"
                    )
                else:
                    st.info("指定した期間にデータがありません。")
            else:
                monthly_chart = build_time_series_chart(
                    chart_filtered, ts_metric_label
                ).properties(height=chart_height)
                st.altair_chart(
                    cast(alt.Chart, monthly_chart), use_container_width=True
                )
        else:
            annual_metric_label = st.radio(
                "指標",
                list(ANNUAL_METRICS.keys()),
                horizontal=True,
                key="annual_metric_export",
            )
            selected_years_ui: list[int] = []
            if not chart_year_options:
                st.info(
                    "選択中の値の種類では、年別同月比較に使えるデータがありません。"
                )
            else:
                selected_years_ui = st.multiselect(
                    "年（同月比較に使う年）",
                    options=chart_year_options,
                    default=selected_years_for_export,
                    key="annual_years_export",
                )
                selected_years_for_export = normalize_selected_years(
                    selected_years_ui, chart_year_options
                )

            if not chart_year_options:
                pass
            elif not selected_years_ui:
                st.info("年を1つ以上選択してください。")
            else:
                yearly_chart = build_yearly_month_compare_chart(
                    chart_scope_all,
                    ANNUAL_METRICS[annual_metric_label],
                    selected_years_for_export,
                ).properties(height=chart_height)
                st.altair_chart(yearly_chart, use_container_width=True)

    if show_mode == "表＋グラフ":
        st.markdown("")

    if show_mode in ["表＋グラフ", "表のみ"]:
        st.subheader("表")
        st.dataframe(table, use_container_width=True, hide_index=True, height=560)
        excel_report_bytes = build_excel_report_bytes(
            d,
            chart_filtered,
            chart_scope_all,
            {
                "scope_type": scope_type,
                "scope_id": scope_id,
                "ym_from": ym_from,
                "ym_to": ym_to,
                "time_series_label": ts_metric_label,
                "annual_metric_label": annual_metric_label,
                "annual_years": selected_years_for_export,
                "chart_value_mode_label": chart_value_mode_label,
            },
        )
        st.download_button(
            "Excelダウンロード（データ＋グラフ）",
            data=excel_report_bytes,
            file_name=f"{export_file_stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()
    st.caption("出典：観光庁『宿泊旅行統計調査』（推移表Excelを取得して整形）")
    st.caption(
        "データは毎週自動更新です。取得元サイトの構造変更等により更新が遅れる場合があります。"
    )


TCD_META_PATH = DATA_DIR / "meta_tcd.json"
TCD_TABLE_NAME = "tcd_stay_nights"

RELEASE_FINAL = "確報"
RELEASE_SECOND_PRELIM = "2次速報"

TCD_SEGMENT_LABELS = {
    "domestic_total": "\u56fd\u5185\u65c5\u884c\uff08\u5408\u8a08\uff09",
    "domestic_business": "\u56fd\u5185\u65c5\u884c\uff08\u51fa\u5f35\u30fb\u696d\u52d9\uff09",
}
TCD_PERIOD_TYPE_LABELS = {
    "annual": "\u5e74\u6b21",
    "quarter": "\u56db\u534a\u671f",
}
TCD_PERIOD_TYPE_KEYS = {v: k for k, v in TCD_PERIOD_TYPE_LABELS.items()}
TCD_RELEASE_FILTERS = {
    "\u6700\u65b0\uff08\u78ba\u5831\u512a\u5148\uff09": "latest",
    "\u78ba\u5831": RELEASE_FINAL,
    "2\u6b21\u901f\u5831": RELEASE_SECOND_PRELIM,
}
TCD_NIGHTS_BIN_ORDER = [
    "1泊",
    "2泊",
    "3泊",
    "4泊",
    "5泊",
    "6泊",
    "7泊",
    "8泊以上",
]


def load_tcd_meta() -> dict:
    if not TCD_META_PATH.exists():
        return {}
    return json.loads(TCD_META_PATH.read_text(encoding="utf-8"))


@st.cache_data(show_spinner=False)
def load_tcd_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {TCD_TABLE_NAME}", conn)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    str_cols = [
        "period_type",
        "period_key",
        "period_label",
        "release_type",
        "segment",
        "nights_bin",
        "source_url",
        "source_title",
    ]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str)

    if "value" in df.columns:
        df["value"] = pd.to_numeric(df["value"], errors="coerce")

    return df


def parse_tcd_period_sort_key(period_key: str) -> tuple[int, int]:
    m_quarter = re.fullmatch(r"(\d{4})Q([1-4])", str(period_key))
    if m_quarter:
        return int(m_quarter.group(1)), int(m_quarter.group(2))

    m_annual = re.fullmatch(r"(\d{4})", str(period_key))
    if m_annual:
        return int(m_annual.group(1)), 0

    return 0, 0


def get_tcd_period_label_map(df: pd.DataFrame) -> dict[str, str]:
    if df.empty:
        return {}

    label_map: dict[str, str] = {}
    for _, row in df[["period_key", "period_label"]].drop_duplicates().iterrows():
        key = str(row["period_key"])
        label = str(row["period_label"]) if pd.notna(row["period_label"]) else key
        label_map[key] = label
    return label_map


def resolve_tcd_latest_period_rows(
    df_for_period_type: pd.DataFrame,
) -> tuple[pd.DataFrame, str | None, str | None]:
    if df_for_period_type.empty:
        return pd.DataFrame(), None, None

    period_keys = sorted(
        df_for_period_type["period_key"].dropna().astype(str).unique().tolist(),
        key=parse_tcd_period_sort_key,
    )
    if not period_keys:
        return pd.DataFrame(), None, None

    latest_period_key = period_keys[-1]
    latest_rows = df_for_period_type[
        df_for_period_type["period_key"] == latest_period_key
    ].copy()

    if latest_rows.empty:
        return pd.DataFrame(), None, None

    if (latest_rows["release_type"] == RELEASE_FINAL).any():
        return (
            latest_rows[latest_rows["release_type"] == RELEASE_FINAL].copy(),
            latest_period_key,
            RELEASE_FINAL,
        )

    if (latest_rows["release_type"] == RELEASE_SECOND_PRELIM).any():
        return (
            latest_rows[latest_rows["release_type"] == RELEASE_SECOND_PRELIM].copy(),
            latest_period_key,
            RELEASE_SECOND_PRELIM,
        )

    release_type = str(latest_rows["release_type"].iloc[0])
    return latest_rows, latest_period_key, release_type


def estimate_tcd_los_by_segment(
    df_period: pd.DataFrame, upper_open_bin_nights: float = 8.5
) -> pd.DataFrame:
    """
    Approximate LOS (average length of stay) from nights-bin total nights.
    8泊以上 is treated as `upper_open_bin_nights`.
    """
    representative_nights = {
        "1\u6cca": 1.0,
        "2\u6cca": 2.0,
        "3\u6cca": 3.0,
        "4\u6cca": 4.0,
        "5\u6cca": 5.0,
        "6\u6cca": 6.0,
        "7\u6cca": 7.0,
        "8\u6cca\u4ee5\u4e0a": float(upper_open_bin_nights),
    }

    work = (
        df_period.groupby(["segment", "nights_bin"], as_index=False)["value"]
        .sum()
        .copy()
    )
    work["rep_nights"] = pd.to_numeric(
        work["nights_bin"].map(representative_nights), errors="coerce"
    )
    work = work[work["rep_nights"].notna()].copy()
    if work.empty:
        return pd.DataFrame()

    work["estimated_stays"] = work["value"] / work["rep_nights"]
    summary = (
        work.groupby("segment", as_index=False)
        .agg(total_nights=("value", "sum"), estimated_stays=("estimated_stays", "sum"))
        .copy()
    )
    summary = summary[summary["estimated_stays"] > 0].copy()
    if summary.empty:
        return pd.DataFrame()

    one_night = (
        work[work["nights_bin"] == "1\u6cca"]
        .groupby("segment", as_index=False)["estimated_stays"]
        .sum()
        .rename(columns={"estimated_stays": "one_night_stays"})
    )
    summary = summary.merge(one_night, on="segment", how="left")
    summary["one_night_stays"] = summary["one_night_stays"].fillna(0.0)
    summary["two_plus_stays"] = (
        summary["estimated_stays"] - summary["one_night_stays"]
    ).clip(lower=0.0)

    summary["estimated_los"] = summary["total_nights"] / summary["estimated_stays"]
    summary["one_night_share_pct"] = (
        summary["one_night_stays"] / summary["estimated_stays"] * 100.0
    )
    summary["two_plus_share_pct"] = (
        summary["two_plus_stays"] / summary["estimated_stays"] * 100.0
    )
    summary["segment_label"] = summary["segment"].map(TCD_SEGMENT_LABELS)

    segment_order = list(TCD_SEGMENT_LABELS.keys())
    summary["segment_order"] = summary["segment"].map(
        {segment: idx for idx, segment in enumerate(segment_order)}
    )
    summary = summary.sort_values("segment_order").drop(columns=["segment_order"])
    return summary.reset_index(drop=True)


def build_tcd_chart(df_period: pd.DataFrame) -> alt.Chart:
    chart_df = (
        df_period.groupby(["nights_bin", "segment"], as_index=False)["value"]
        .sum()
        .copy()
    )
    chart_df["segment_label"] = chart_df["segment"].map(TCD_SEGMENT_LABELS)

    available_bins = [
        b
        for b in TCD_NIGHTS_BIN_ORDER
        if b in chart_df["nights_bin"].astype(str).tolist()
    ]

    return (
        alt.Chart(chart_df)
        .mark_bar()
        .encode(
            x=alt.X(
                "nights_bin:N",
                title="\u5bbf\u6cca\u6570\u533a\u5206",
                sort=available_bins,
            ),
            xOffset=alt.XOffset(
                "segment_label:N", sort=list(TCD_SEGMENT_LABELS.values())
            ),
            y=alt.Y("value:Q", title="\u5ef6\u3079\u6cca\u6570"),
            color=alt.Color(
                "segment_label:N",
                title="\u7cfb\u5217",
                sort=list(TCD_SEGMENT_LABELS.values()),
            ),
            tooltip=[
                alt.Tooltip("nights_bin:N", title="\u5bbf\u6cca\u6570\u533a\u5206"),
                alt.Tooltip("segment_label:N", title="\u7cfb\u5217"),
                alt.Tooltip("value:Q", title="\u5ef6\u3079\u6cca\u6570", format=",.0f"),
            ],
        )
    )


def render_tcd_view() -> None:
    st.title(
        "\u65c5\u884c\u30fb\u89b3\u5149\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb\uff1a"
        "\u5bbf\u6cca\u6570(8\u533a\u5206)\u5225 \u5ef6\u3079\u6cca\u6570\uff08\u5168\u56fd\uff09"
    )

    meta = load_tcd_meta()
    if meta:
        st.caption(
            f"\u6700\u7d42\u78ba\u8a8d\uff08UTC\uff09: {meta.get('last_checked_at')} / "
            f"\u51e6\u7406\u6e08\u307f\u30d5\u30a1\u30a4\u30eb\u6570: {len(meta.get('processed_files', []))}"
        )

    df = load_tcd_data()
    if df.empty:
        st.error(
            "TCD\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
            "\u5148\u306b `python -m scripts.update_tcd_data` \u3092\u5b9f\u884c\u3057\u3066"
            " data/ \u3092\u751f\u6210\u3057\u3066\u304f\u3060\u3055\u3044\u3002"
        )
        return

    work = df[df["segment"].isin(TCD_SEGMENT_LABELS.keys())].copy()
    if work.empty:
        st.error(
            "\u8868\u793a\u306b\u5fc5\u8981\u306a\u7cfb\u5217"
            "\uff08domestic_total / domestic_business\uff09\u304c\u898b\u3064\u304b\u308a\u307e\u305b\u3093\u3002"
        )
        return

    col1, col2, col3 = st.columns([2, 2, 4])
    with col1:
        period_type_label = st.radio(
            "\u671f\u9593\u7a2e\u5225",
            list(TCD_PERIOD_TYPE_KEYS.keys()),
            horizontal=True,
            key="tcd_period_type",
        )
    period_type = TCD_PERIOD_TYPE_KEYS[period_type_label]
    work = work[work["period_type"] == period_type].copy()

    if work.empty:
        st.warning(
            "\u9078\u629e\u3057\u305f\u671f\u9593\u7a2e\u5225\u306e\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    with col2:
        release_filter_label = st.radio(
            "\u30ea\u30ea\u30fc\u30b9\u7a2e\u5225",
            list(TCD_RELEASE_FILTERS.keys()),
            key="tcd_release_filter",
        )
    release_filter = TCD_RELEASE_FILTERS[release_filter_label]

    selected_period_key: str | None = None
    selected_release_type: str | None = None

    if release_filter == "latest":
        filtered, selected_period_key, selected_release_type = (
            resolve_tcd_latest_period_rows(work)
        )
    else:
        filtered_by_release = work[work["release_type"] == release_filter].copy()
        if filtered_by_release.empty:
            st.warning(
                "\u9078\u629e\u3057\u305f\u30ea\u30ea\u30fc\u30b9\u7a2e\u5225\u306e\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
            )
            return

        period_options = sorted(
            filtered_by_release["period_key"].dropna().astype(str).unique().tolist(),
            key=parse_tcd_period_sort_key,
            reverse=True,
        )
        period_label_map = get_tcd_period_label_map(filtered_by_release)

        with col3:
            selected_period_key = st.selectbox(
                "\u8868\u793a\u671f\u9593",
                period_options,
                format_func=lambda k: f"{period_label_map.get(k, k)} ({k})",
                key=f"tcd_period_key_{period_type}_{release_filter}",
            )

        filtered = filtered_by_release[
            filtered_by_release["period_key"] == selected_period_key
        ].copy()
        selected_release_type = release_filter

    if filtered.empty or selected_period_key is None or selected_release_type is None:
        st.warning(
            "\u8868\u793a\u5bfe\u8c61\u306e\u30c7\u30fc\u30bf\u304c\u898b\u3064\u304b\u308a\u307e\u305b\u3093\u3002"
        )
        return

    period_label_map = get_tcd_period_label_map(filtered)
    period_label = period_label_map.get(selected_period_key, selected_period_key)
    st.caption(f"\u8868\u793a\u4e2d: {period_label} / {selected_release_type}")

    los_open_bin = st.number_input(
        "LOS\u6982\u7b97: 8\u6cca\u4ee5\u4e0a\u306e\u4ee3\u8868\u5024\uff08\u6cca\uff09",
        min_value=8.0,
        max_value=15.0,
        value=8.5,
        step=0.5,
        key="tcd_los_open_bin",
    )
    los_summary = estimate_tcd_los_by_segment(
        filtered, upper_open_bin_nights=los_open_bin
    )
    if not los_summary.empty:
        st.subheader("LOS\uff08\u5e73\u5747\u6cca\u6570\uff09\u6982\u7b97")
        metric_cols = st.columns(len(los_summary))
        for col, row in zip(metric_cols, los_summary.itertuples(index=False)):
            col.metric(str(row.segment_label), f"{row.estimated_los:.2f} \u6cca")
        st.caption(
            "\u8a08\u7b97\u5f0f: LOS \u2248 \u03a3(\u5ef6\u3079\u6cca\u6570) / "
            "\u03a3(\u5404\u533a\u5206\u306e\u5ef6\u3079\u6cca\u6570 / \u533a\u5206\u4ee3\u8868\u5024)"
        )

        share_display = (
            los_summary[["segment_label", "one_night_share_pct", "two_plus_share_pct"]]
            .rename(
                columns={
                    "segment_label": "\u7cfb\u5217",
                    "one_night_share_pct": "1\u6cca\u30b7\u30a7\u30a2",
                    "two_plus_share_pct": "2\u6cca\u4ee5\u4e0a\u30b7\u30a7\u30a2",
                }
            )
            .copy()
        )
        share_display["1\u6cca\u30b7\u30a7\u30a2"] = share_display[
            "1\u6cca\u30b7\u30a7\u30a2"
        ].map(lambda v: f"{v:.1f}%")
        share_display["2\u6cca\u4ee5\u4e0a\u30b7\u30a7\u30a2"] = share_display[
            "2\u6cca\u4ee5\u4e0a\u30b7\u30a7\u30a2"
        ].map(lambda v: f"{v:.1f}%")
        st.subheader(
            "\u5bbf\u6cca\u65e5\u6570\u30b7\u30a7\u30a2\uff08\u63a8\u5b9a\u4ef6\u6570\u30d9\u30fc\u30b9\uff09"
        )
        st.dataframe(share_display, use_container_width=True, hide_index=True)
        st.caption(
            "\u8a08\u7b97\u5f0f: \u30b7\u30a7\u30a2 \u2248 "
            "\u03a3(\u533a\u5206\u306e\u5ef6\u3079\u6cca\u6570 / \u533a\u5206\u4ee3\u8868\u5024) / "
            "\u03a3(\u5168\u533a\u5206\u306e\u5ef6\u3079\u6cca\u6570 / \u533a\u5206\u4ee3\u8868\u5024)"
        )

    chart = build_tcd_chart(filtered).properties(height=500)
    st.altair_chart(chart, use_container_width=True)

    table_df = (
        filtered.groupby(["nights_bin", "segment"], as_index=False)["value"]
        .sum()
        .copy()
    )
    table_df["segment_label"] = table_df["segment"].map(TCD_SEGMENT_LABELS)
    table_pivot = (
        table_df.pivot(index="nights_bin", columns="segment_label", values="value")
        .reindex(TCD_NIGHTS_BIN_ORDER)
        .reset_index()
    )
    table_pivot = table_pivot.rename(
        columns={"nights_bin": "\u5bbf\u6cca\u6570\u533a\u5206"}
    )

    st.subheader("\u8868")
    st.dataframe(table_pivot, use_container_width=True, hide_index=True)
    st.caption(
        "\u30c7\u30fc\u30bf\u306f\u6bce\u9031\u81ea\u52d5\u66f4\u65b0\u3067\u3059\u3002"
        "\u53d6\u5f97\u5143\u30b5\u30a4\u30c8\u306e\u69cb\u9020\u5909\u66f4\u7b49\u306b\u3088\u308a"
        "\u66f4\u65b0\u304c\u9045\u308c\u308b\u5834\u5408\u304c\u3042\u308a\u307e\u3059\u3002"
    )


ICD_META_PATH = DATA_DIR / "meta_icd.json"
ICD_SPEND_TABLE_NAME = "icd_spend_items"
ICD_ENTRY_TABLE_NAME = "icd_entry_port_summary"

TA_META_PATH = DATA_DIR / "meta_ta.json"
TA_TABLE_NAME = "ta_company_amounts"

AIRPORT_VOLUME_META_PATH = DATA_DIR / "meta_airport_volume.json"
AIRPORT_VOLUME_TABLE_NAME = "airport_arrivals_monthly"

DATASET_LABEL_STAY = "\u5bbf\u6cca\u65c5\u884c\u7d71\u8a08\u8abf\u67fb"
DATASET_LABEL_TCD = "\u65c5\u884c\u30fb\u89b3\u5149\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb"
DATASET_LABEL_ICD = (
    "\u30a4\u30f3\u30d0\u30a6\u30f3\u30c9\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb"
)
DATASET_LABEL_TA = "\u65c5\u884c\u696d\u8005\u53d6\u6271\u984d"
DATASET_LABEL_AIRPORT_VOLUME = "\u7a7a\u6e2f\u5225\u5165\u56fd\u8005\u6570"
DATASET_LABEL_EVENTS_OFFICIAL = "全国イベント情報（会場公式）"
DATASET_LABEL_EVENTS_SIGNALS = "全国イベント速報（ニュース）"
DATASET_LABEL_EVENTS = DATASET_LABEL_EVENTS_OFFICIAL

EVENTS_DB_PATH = DATA_DIR / "events.sqlite"
EVENT_SIGNALS_DB_PATH = DATA_DIR / "event_signals.sqlite"
EVENTS_ARTIST_INFERRED_PATH = DATA_DIR / "events_artist_inferred.csv"

ICD_PURPOSE_LABELS = {
    "all": "\u5168\u76ee\u7684",
    "leisure": "\u89b3\u5149\u30fb\u30ec\u30b8\u30e3\u30fc",
}
ICD_PURPOSE_KEYS = {v: k for k, v in ICD_PURPOSE_LABELS.items()}
ICD_PORT_TYPE_LABELS = {
    "entry": "\u5165\u56fd",
    "exit": "\u51fa\u56fd",
}
ICD_PORT_TYPE_KEYS = {v: k for k, v in ICD_PORT_TYPE_LABELS.items()}

TA_SEGMENT_LABELS = {
    "overseas": "\u6d77\u5916\u65c5\u884c",
    "foreign": "\u5916\u56fd\u4eba\u65c5\u884c",
    "domestic": "\u56fd\u5185\u65c5\u884c",
    "total": "\u5408\u8a08",
}
TA_SEGMENT_KEYS = {v: k for k, v in TA_SEGMENT_LABELS.items()}


def parse_period_sort_key(period_key: str) -> tuple[int, int, int]:
    m_quarter = re.fullmatch(r"(\d{4})Q([1-4])", str(period_key))
    if m_quarter:
        return int(m_quarter.group(1)), int(m_quarter.group(2)), 0

    m_month = re.fullmatch(r"(\d{4})-(\d{2})", str(period_key))
    if m_month:
        return int(m_month.group(1)), 9, int(m_month.group(2))

    m_annual = re.fullmatch(r"(\d{4})", str(period_key))
    if m_annual:
        return int(m_annual.group(1)), 0, 0

    if str(period_key) == "total":
        return 9999, 99, 99

    return 0, 0, 0


def format_metric_value(value: float | None, digits: int = 0, suffix: str = "") -> str:
    if value is None or pd.isna(value):
        return "\u2014"
    return f"{float(value):,.{digits}f}{suffix}"


def first_numeric_value(df: pd.DataFrame, column: str) -> float | None:
    if df.empty or column not in df.columns:
        return None
    series = pd.to_numeric(df[column], errors="coerce").dropna()
    if series.empty:
        return None
    return float(series.iloc[0])


def load_icd_meta() -> dict:
    if not ICD_META_PATH.exists():
        return {}
    return json.loads(ICD_META_PATH.read_text(encoding="utf-8"))


@st.cache_data(show_spinner=False)
def load_icd_spend_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {ICD_SPEND_TABLE_NAME}", conn)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    str_cols = [
        "period_label",
        "period_key",
        "release_type",
        "purpose",
        "nationality",
        "item_group",
        "item",
    ]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str)

    for col in ["spend_yen", "share_pct"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


@st.cache_data(show_spinner=False)
def load_icd_entry_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {ICD_ENTRY_TABLE_NAME}", conn)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    str_cols = [
        "period_label",
        "period_key",
        "release_type",
        "purpose",
        "port_type",
        "entry_port",
        "nationality",
    ]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str)

    for col in ["respondents", "spend_yen", "avg_nights", "spend_per_night_yen"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def map_icd_item_bucket(item_name: str) -> str:
    item = str(item_name or "")
    if "\u5bbf\u6cca" in item:
        return "\u5bbf\u6cca"
    if "\u98f2\u98df" in item:
        return "\u98f2\u98df"
    if any(
        keyword in item
        for keyword in [
            "\u4ea4\u901a",
            "\u9244\u9053",
            "\u30d0\u30b9",
            "\u30bf\u30af\u30b7\u30fc",
            "\u30ec\u30f3\u30bf\u30ab\u30fc",
            "\u822a\u7a7a",
            "\u8239\u8236",
            "Rail Pass",
        ]
    ):
        return "\u4ea4\u901a"
    if any(
        keyword in item
        for keyword in [
            "\u8cb7\u7269",
            "\u83d3\u5b50",
            "\u9152",
            "\u98df\u6599",
            "\u5316\u7ca7",
            "\u533b\u85ac",
            "\u5065\u5eb7",
            "\u8863\u985e",
            "\u9774",
            "\u304b\u3070\u3093",
            "\u9769\u88fd\u54c1",
            "\u96fb\u6c17\u88fd\u54c1",
            "\u6642\u8a08",
            "\u5b9d\u77f3",
            "\u6c11\u82b8\u54c1",
        ]
    ):
        return "\u8cb7\u7269"
    if any(
        keyword in item
        for keyword in [
            "\u5a2f\u697d",
            "\u30c4\u30a2\u30fc",
            "\u30c6\u30fc\u30de\u30d1\u30fc\u30af",
            "\u821e\u53f0",
            "\u97f3\u697d",
            "\u30b9\u30dd\u30fc\u30c4",
            "\u7f8e\u8853\u9928",
            "\u535a\u7269\u9928",
            "\u6e29\u6cc9",
            "\u30de\u30c3\u30b5\u30fc\u30b8",
            "\u30b3\u30f3\u30d9\u30f3\u30b7\u30e7\u30f3",
        ]
    ):
        return "\u5a2f\u697d"
    return "\u305d\u306e\u4ed6"


def render_icd_view() -> None:
    st.title(
        "\u30a4\u30f3\u30d0\u30a6\u30f3\u30c9\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb\uff1a"
        "\u56fd\u7c4d\u5225\u8cbb\u76ee\u69cb\u6210\u3068\u5165\u56fd\u7a7a\u6e2f\u5225\u6982\u6cc1"
    )

    meta = load_icd_meta()
    if meta:
        st.caption(
            f"\u6700\u7d42\u78ba\u8a8d\uff08UTC\uff09: {meta.get('fetched_at_utc')} / "
            f"\u5bfe\u8c61: {meta.get('period_label', '-')}"
            f" {meta.get('release_type', '')}"
        )

    spend_df = load_icd_spend_data()
    entry_df = load_icd_entry_data()
    if spend_df.empty and entry_df.empty:
        st.error(
            "ICD\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
            "\u5148\u306b `python -m scripts.update_icd_data` \u3092\u5b9f\u884c\u3057\u3066"
            " data/ \u3092\u751f\u6210\u3057\u3066\u304f\u3060\u3055\u3044\u3002"
        )
        return

    period_records: list[pd.DataFrame] = []
    for df in [spend_df, entry_df]:
        if not df.empty and {"period_key", "period_label"}.issubset(df.columns):
            period_records.append(df[["period_key", "period_label"]].drop_duplicates())

    if not period_records:
        st.warning(
            "\u671f\u9593\u9078\u629e\u7528\u306e\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    periods = pd.concat(period_records, ignore_index=True).drop_duplicates()
    period_label_map = {
        str(row["period_key"]): str(row["period_label"])
        for _, row in periods.iterrows()
    }
    period_options = sorted(
        periods["period_key"].astype(str).unique().tolist(),
        key=parse_period_sort_key,
        reverse=True,
    )
    if not period_options:
        st.warning(
            "\u8868\u793a\u53ef\u80fd\u306a\u671f\u9593\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    col1, col2, col3 = st.columns([3, 2, 3])
    with col1:
        selected_period_key = st.selectbox(
            "\u671f\u9593",
            options=period_options,
            index=0,
            format_func=lambda k: f"{period_label_map.get(k, k)} ({k})",
            key="icd_period_key",
        )

    available_purposes = sorted(
        {
            *spend_df.get("purpose", pd.Series(dtype=str))
            .astype(str)
            .unique()
            .tolist(),
            *entry_df.get("purpose", pd.Series(dtype=str))
            .astype(str)
            .unique()
            .tolist(),
        }
    )
    available_purpose_keys = [p for p in ICD_PURPOSE_LABELS if p in available_purposes]
    if not available_purpose_keys:
        available_purpose_keys = ["all"]

    with col2:
        purpose_label = st.radio(
            "\u76ee\u7684",
            [ICD_PURPOSE_LABELS[p] for p in available_purpose_keys],
            horizontal=True,
            key="icd_purpose",
        )
    selected_purpose = ICD_PURPOSE_KEYS[purpose_label]

    spend_filtered = spend_df[
        (spend_df["period_key"] == selected_period_key)
        & (spend_df["purpose"] == selected_purpose)
    ].copy()
    entry_filtered = entry_df[
        (entry_df["period_key"] == selected_period_key)
        & (entry_df["purpose"] == selected_purpose)
    ].copy()
    if not entry_filtered.empty and "port_type" not in entry_filtered.columns:
        entry_filtered["port_type"] = "entry"

    nationality_options = sorted(
        {
            *spend_filtered.get("nationality", pd.Series(dtype=str))
            .astype(str)
            .unique()
            .tolist(),
            *entry_filtered.get("nationality", pd.Series(dtype=str))
            .astype(str)
            .unique()
            .tolist(),
        }
    )
    if "\u5168\u56fd\u7c4d\uff65\u5730\u57df" in nationality_options:
        nationality_options = [
            "\u5168\u56fd\u7c4d\uff65\u5730\u57df",
            *[
                n
                for n in nationality_options
                if n != "\u5168\u56fd\u7c4d\uff65\u5730\u57df"
            ],
        ]
    if not nationality_options:
        st.warning(
            "\u9078\u629e\u6761\u4ef6\u306e\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    with col3:
        selected_nationality = st.selectbox(
            "\u56fd\u7c4d\u30fb\u5730\u57df",
            options=nationality_options,
            key="icd_nationality",
        )

    available_port_type_keys = (
        [
            k
            for k in ICD_PORT_TYPE_LABELS
            if not entry_filtered.empty
            and k in entry_filtered["port_type"].astype(str).unique().tolist()
        ]
        if not entry_filtered.empty
        else []
    )
    if not available_port_type_keys:
        available_port_type_keys = ["entry"]
    selected_port_type = available_port_type_keys[0]
    stored_port_type_label = st.session_state.get("icd_port_type")
    if stored_port_type_label in ICD_PORT_TYPE_KEYS:
        stored_port_type_key = ICD_PORT_TYPE_KEYS[stored_port_type_label]
        if stored_port_type_key in available_port_type_keys:
            selected_port_type = stored_port_type_key

    total_row = entry_filtered[
        (entry_filtered["port_type"] == selected_port_type)
        & (entry_filtered["entry_port"] == "\u5168\u4f53")
        & (entry_filtered["nationality"] == selected_nationality)
    ].copy()
    total_spend = first_numeric_value(total_row, "spend_yen")
    avg_nights = first_numeric_value(total_row, "avg_nights")

    lodging_row = spend_filtered[
        (spend_filtered["nationality"] == selected_nationality)
        & (spend_filtered["item"] == "\u5bbf\u6cca\u8cbb")
    ].copy()
    lodging_spend = first_numeric_value(lodging_row, "spend_yen")
    lodging_per_night = (
        lodging_spend / avg_nights
        if lodging_spend is not None and avg_nights is not None and avg_nights > 0
        else None
    )

    kpi_cols = st.columns(4)
    kpi_cols[0].metric(
        "\u7dcf\u652f\u51fa\uff08\u5186/\u4eba\uff09", format_metric_value(total_spend)
    )
    kpi_cols[1].metric(
        "\u5e73\u5747\u6cca\u6570\uff08\u6cca\uff09", format_metric_value(avg_nights, 2)
    )
    kpi_cols[2].metric(
        "\u5bbf\u6cca\u8cbb\uff08\u5186/\u4eba\uff09",
        format_metric_value(lodging_spend),
    )
    kpi_cols[3].metric(
        "\u5bbf\u6cca\u8cbb/\u6cca\uff08\u5186\uff09",
        format_metric_value(lodging_per_night),
    )

    st.subheader("\u56fd\u7c4d\u5225 TopN")
    top_n = st.slider("TopN", min_value=5, max_value=20, value=10, key="icd_top_n")
    nationality_totals = entry_filtered[
        (entry_filtered["port_type"] == selected_port_type)
        & (entry_filtered["entry_port"] == "\u5168\u4f53")
    ].copy()

    col_a, col_b = st.columns(2)
    with col_a:
        top_spend = (
            nationality_totals.dropna(subset=["spend_yen"])
            .sort_values("spend_yen", ascending=False)
            .head(top_n)
        )
        if top_spend.empty:
            st.info(
                "\u7dcf\u652f\u51faTopN\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
            )
        else:
            chart_spend = (
                alt.Chart(top_spend)
                .mark_bar()
                .encode(
                    x=alt.X(
                        "spend_yen:Q",
                        title="\u7dcf\u652f\u51fa\uff08\u5186/\u4eba\uff09",
                    ),
                    y=alt.Y(
                        "nationality:N",
                        sort="-x",
                        title="\u56fd\u7c4d\u30fb\u5730\u57df",
                    ),
                    tooltip=[
                        alt.Tooltip(
                            "nationality:N", title="\u56fd\u7c4d\u30fb\u5730\u57df"
                        ),
                        alt.Tooltip(
                            "spend_yen:Q", title="\u7dcf\u652f\u51fa", format=",.0f"
                        ),
                    ],
                )
                .properties(height=380)
            )
            st.altair_chart(chart_spend, use_container_width=True)

    with col_b:
        top_nights = (
            nationality_totals.dropna(subset=["avg_nights"])
            .sort_values("avg_nights", ascending=False)
            .head(top_n)
        )
        if top_nights.empty:
            st.info(
                "\u5e73\u5747\u6cca\u6570TopN\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
            )
        else:
            chart_nights = (
                alt.Chart(top_nights)
                .mark_bar(color="#4C78A8")
                .encode(
                    x=alt.X(
                        "avg_nights:Q",
                        title="\u5e73\u5747\u6cca\u6570\uff08\u6cca\uff09",
                    ),
                    y=alt.Y(
                        "nationality:N",
                        sort="-x",
                        title="\u56fd\u7c4d\u30fb\u5730\u57df",
                    ),
                    tooltip=[
                        alt.Tooltip(
                            "nationality:N", title="\u56fd\u7c4d\u30fb\u5730\u57df"
                        ),
                        alt.Tooltip(
                            "avg_nights:Q",
                            title="\u5e73\u5747\u6cca\u6570",
                            format=".2f",
                        ),
                    ],
                )
                .properties(height=380)
            )
            st.altair_chart(chart_nights, use_container_width=True)

    st.subheader("\u56fd\u7c4d\u5225 \u8cbb\u76ee\u69cb\u6210")
    if spend_filtered.empty:
        st.info(
            "\u8cbb\u76ee\u69cb\u6210\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        top_nat_list = (
            top_spend["nationality"].astype(str).tolist() if not top_spend.empty else []
        )
        comp_source = spend_filtered.copy()
        if top_nat_list:
            comp_source = comp_source[
                comp_source["nationality"].isin(top_nat_list)
            ].copy()
        comp_source["bucket"] = comp_source["item"].map(map_icd_item_bucket)
        comp_summary = (
            comp_source.groupby(["nationality", "bucket"], as_index=False)["spend_yen"]
            .sum(min_count=1)
            .dropna(subset=["spend_yen"])
        )
        if comp_summary.empty:
            st.info(
                "\u8cbb\u76ee\u69cb\u6210\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
            )
        else:
            bucket_order = [
                "\u5bbf\u6cca",
                "\u98f2\u98df",
                "\u4ea4\u901a",
                "\u8cb7\u7269",
                "\u5a2f\u697d",
                "\u305d\u306e\u4ed6",
            ]
            chart_comp = (
                alt.Chart(comp_summary)
                .mark_bar()
                .encode(
                    x=alt.X("nationality:N", title="\u56fd\u7c4d\u30fb\u5730\u57df"),
                    y=alt.Y(
                        "spend_yen:Q",
                        title="\u8cbb\u76ee\u5225\u652f\u51fa\uff08\u5186/\u4eba\uff09",
                    ),
                    color=alt.Color(
                        "bucket:N", title="\u8cbb\u76ee", sort=bucket_order
                    ),
                    tooltip=[
                        alt.Tooltip(
                            "nationality:N", title="\u56fd\u7c4d\u30fb\u5730\u57df"
                        ),
                        alt.Tooltip("bucket:N", title="\u8cbb\u76ee"),
                        alt.Tooltip("spend_yen:Q", title="\u652f\u51fa", format=",.0f"),
                    ],
                )
                .properties(height=420)
            )
            st.altair_chart(chart_comp, use_container_width=True)

    port_section_label = (
        "\u5165\u56fd\u7a7a\u6e2f\u30fb\u6e2f\u5225"
        if selected_port_type == "entry"
        else "\u51fa\u56fd\u7a7a\u6e2f\u30fb\u6e2f\u5225"
    )
    selected_port_type_label = st.radio(
        "\u6e2f\u533a\u5206",
        [ICD_PORT_TYPE_LABELS[k] for k in available_port_type_keys],
        index=available_port_type_keys.index(selected_port_type),
        horizontal=True,
        key="icd_port_type",
    )
    selected_port_type = ICD_PORT_TYPE_KEYS[selected_port_type_label]
    port_section_label = (
        "\u5165\u56fd\u7a7a\u6e2f\u30fb\u6e2f\u5225"
        if selected_port_type == "entry"
        else "\u51fa\u56fd\u7a7a\u6e2f\u30fb\u6e2f\u5225"
    )
    st.subheader(port_section_label)
    entry_nat = entry_filtered[
        (entry_filtered["port_type"] == selected_port_type)
        & (entry_filtered["nationality"] == selected_nationality)
        & (entry_filtered["entry_port"] != "\u5168\u4f53")
    ].copy()
    if entry_nat.empty:
        st.info(
            f"{port_section_label}\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        col_c, col_d = st.columns(2)
        with col_c:
            top_ports = (
                entry_nat.dropna(subset=["respondents"])
                .sort_values("respondents", ascending=False)
                .head(top_n)
            )
            if top_ports.empty:
                st.info(
                    "\u56de\u7b54\u8005\u6570TopN\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
                )
            else:
                chart_ports = (
                    alt.Chart(top_ports)
                    .mark_bar(color="#72B7B2")
                    .encode(
                        x=alt.X("respondents:Q", title="\u56de\u7b54\u8005\u6570"),
                        y=alt.Y("entry_port:N", sort="-x", title=port_section_label),
                        tooltip=[
                            alt.Tooltip("entry_port:N", title=port_section_label),
                            alt.Tooltip(
                                "respondents:Q",
                                title="\u56de\u7b54\u8005\u6570",
                                format=",.0f",
                            ),
                        ],
                    )
                    .properties(height=420)
                )
                st.altair_chart(chart_ports, use_container_width=True)

        with col_d:
            scatter_base = entry_nat.dropna(subset=["spend_yen", "avg_nights"]).copy()
            if scatter_base.empty:
                st.info(
                    "\u7dcf\u652f\u51fa/\u5e73\u5747\u6cca\u6570\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
                )
            else:
                chart_scatter = (
                    alt.Chart(scatter_base)
                    .mark_circle(opacity=0.8)
                    .encode(
                        x=alt.X(
                            "avg_nights:Q",
                            title="\u5e73\u5747\u6cca\u6570\uff08\u6cca\uff09",
                        ),
                        y=alt.Y(
                            "spend_yen:Q",
                            title="\u7dcf\u652f\u51fa\uff08\u5186/\u4eba\uff09",
                        ),
                        size=alt.Size(
                            "respondents:Q", title="\u56de\u7b54\u8005\u6570"
                        ),
                        color=alt.value("#F58518"),
                        tooltip=[
                            alt.Tooltip("entry_port:N", title=port_section_label),
                            alt.Tooltip(
                                "respondents:Q",
                                title="\u56de\u7b54\u8005\u6570",
                                format=",.0f",
                            ),
                            alt.Tooltip(
                                "avg_nights:Q",
                                title="\u5e73\u5747\u6cca\u6570",
                                format=".2f",
                            ),
                            alt.Tooltip(
                                "spend_yen:Q", title="\u7dcf\u652f\u51fa", format=",.0f"
                            ),
                        ],
                    )
                    .properties(height=420)
                )
                st.altair_chart(chart_scatter, use_container_width=True)

    st.caption(
        "\u51fa\u5178\uff1a\u89b3\u5149\u5e81\u300e\u30a4\u30f3\u30d0\u30a6\u30f3\u30c9\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb"
        "\uff08\u65e7 \u8a2a\u65e5\u5916\u56fd\u4eba\u6d88\u8cbb\u52d5\u5411\u8abf\u67fb\uff09\u300f"
        "\uff08\u96c6\u8a08\u8868Excel\u3092\u53d6\u5f97\u3057\u3066\u6574\u5f62\uff09"
    )
    st.caption(
        "\u30c7\u30fc\u30bf\u306f\u6bce\u9031\u81ea\u52d5\u66f4\u65b0\u3067\u3059\u3002"
        "\u53d6\u5f97\u5143\u30b5\u30a4\u30c8\u306e\u69cb\u9020\u5909\u66f4\u7b49\u306b\u3088\u308a"
        "\u66f4\u65b0\u304c\u9045\u308c\u308b\u5834\u5408\u304c\u3042\u308a\u307e\u3059\u3002"
    )


def load_ta_meta() -> dict:
    if not TA_META_PATH.exists():
        return {}
    return json.loads(TA_META_PATH.read_text(encoding="utf-8"))


@st.cache_data(show_spinner=False)
def load_ta_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {TA_TABLE_NAME}", conn)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    for col in ["fiscal_year", "period", "company", "segment"]:
        if col in df.columns:
            df[col] = df[col].astype(str)
    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce")

    return df


def render_ta_view() -> None:
    st.title(
        "\u65c5\u884c\u696d\u8005\u53d6\u6271\u984d\uff1a\u5404\u793e\u5225\u5185\u8a33"
    )

    meta = load_ta_meta()
    if meta:
        st.caption(
            f"\u6700\u7d42\u78ba\u8a8d\uff08UTC\uff09: {meta.get('last_checked_at')} / "
            f"\u51e6\u7406\u6e08\u307f\u30d5\u30a1\u30a4\u30eb\u6570: {len(meta.get('processed_files', []))}"
        )

    df = load_ta_data()
    if df.empty:
        st.error(
            "TA\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
            "\u5148\u306b `python -m scripts.update_ta_data` \u3092\u5b9f\u884c\u3057\u3066"
            " data/ \u3092\u751f\u6210\u3057\u3066\u304f\u3060\u3055\u3044\u3002"
        )
        return

    fiscal_year_options = sorted(
        df["fiscal_year"].dropna().astype(str).unique().tolist(),
        reverse=True,
    )
    if not fiscal_year_options:
        st.warning(
            "\u5e74\u5ea6\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    col1, col2, col3 = st.columns([2, 2, 2])
    with col1:
        fiscal_year = st.selectbox(
            "\u5e74\u5ea6",
            options=fiscal_year_options,
            index=0,
            key="ta_fiscal_year",
        )

    fy_data = df[df["fiscal_year"] == fiscal_year].copy()
    period_options_raw = fy_data["period"].dropna().astype(str).unique().tolist()
    period_options = sorted(period_options_raw, key=parse_period_sort_key, reverse=True)
    if "total" in period_options:
        period_options = ["total", *[p for p in period_options if p != "total"]]

    with col2:
        period = st.selectbox(
            "\u671f\u9593",
            options=period_options,
            format_func=lambda p: "\u5e74\u5ea6\u7dcf\u8a08" if p == "total" else p,
            key="ta_period",
        )

    with col3:
        segment_label = st.selectbox(
            "\u30bb\u30b0\u30e1\u30f3\u30c8",
            options=[
                "\u5168\u30bb\u30b0\u30e1\u30f3\u30c8",
                *TA_SEGMENT_LABELS.values(),
            ],
            key="ta_segment",
        )

    period_data = fy_data[fy_data["period"] == period].copy()
    if segment_label != "\u5168\u30bb\u30b0\u30e1\u30f3\u30c8":
        segment_key = TA_SEGMENT_KEYS[segment_label]
        period_data = period_data[period_data["segment"] == segment_key].copy()

    total_amount = float(period_data["amount"].sum()) if not period_data.empty else 0.0
    st.metric(
        "\u7dcf\u8a08\uff08\u5343\u5186\uff09",
        format_metric_value(total_amount),
    )

    st.subheader("\u4f1a\u793e\u5225\u30e9\u30f3\u30ad\u30f3\u30b0")
    top_n = st.slider("TopN", min_value=5, max_value=30, value=15, key="ta_top_n")
    ranking = (
        period_data.groupby("company", as_index=False)["amount"]
        .sum()
        .sort_values("amount", ascending=False)
        .head(top_n)
    )
    if ranking.empty:
        st.info(
            "\u30e9\u30f3\u30ad\u30f3\u30b0\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        ranking_chart = (
            alt.Chart(ranking)
            .mark_bar()
            .encode(
                x=alt.X("amount:Q", title="\u53d6\u6271\u984d\uff08\u5343\u5186\uff09"),
                y=alt.Y("company:N", sort="-x", title="\u4f1a\u793e"),
                tooltip=[
                    alt.Tooltip("company:N", title="\u4f1a\u793e"),
                    alt.Tooltip("amount:Q", title="\u53d6\u6271\u984d", format=",.0f"),
                ],
            )
            .properties(height=520)
        )
        st.altair_chart(ranking_chart, use_container_width=True)

    st.subheader(
        "\u30bb\u30b0\u30e1\u30f3\u30c8\u5225\u63a8\u79fb\uff08\u6708\u6b21\uff09"
    )
    monthly = fy_data[fy_data["period"].str.match(r"\d{4}-\d{2}$", na=False)].copy()
    if monthly.empty:
        st.info(
            "\u6708\u6b21\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        monthly_summary = (
            monthly.groupby(["period", "segment"], as_index=False)["amount"]
            .sum()
            .copy()
        )
        monthly_summary["segment_label"] = monthly_summary["segment"].map(
            TA_SEGMENT_LABELS
        )
        if segment_label != "\u5168\u30bb\u30b0\u30e1\u30f3\u30c8":
            monthly_summary = monthly_summary[
                monthly_summary["segment_label"] == segment_label
            ].copy()

        if monthly_summary.empty:
            st.info(
                "\u9078\u629e\u30bb\u30b0\u30e1\u30f3\u30c8\u306e\u6708\u6b21\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
            )
        else:
            month_order = sorted(
                monthly_summary["period"].unique().tolist(),
                key=parse_period_sort_key,
            )
            trend_chart = (
                alt.Chart(monthly_summary)
                .mark_line(point=True)
                .encode(
                    x=alt.X("period:N", sort=month_order, title="\u6708"),
                    y=alt.Y(
                        "amount:Q", title="\u53d6\u6271\u984d\uff08\u5343\u5186\uff09"
                    ),
                    color=alt.Color(
                        "segment_label:N", title="\u30bb\u30b0\u30e1\u30f3\u30c8"
                    ),
                    tooltip=[
                        alt.Tooltip("period:N", title="\u6708"),
                        alt.Tooltip(
                            "segment_label:N", title="\u30bb\u30b0\u30e1\u30f3\u30c8"
                        ),
                        alt.Tooltip(
                            "amount:Q", title="\u53d6\u6271\u984d", format=",.0f"
                        ),
                    ],
                )
                .properties(height=360)
            )
            st.altair_chart(trend_chart, use_container_width=True)

    st.subheader("\u8868")
    table = period_data.copy()
    if not table.empty:
        table["segment"] = table["segment"].map(TA_SEGMENT_LABELS)
    table = table.rename(
        columns={
            "fiscal_year": "\u5e74\u5ea6",
            "period": "\u671f\u9593",
            "company": "\u4f1a\u793e",
            "segment": "\u30bb\u30b0\u30e1\u30f3\u30c8",
            "amount": "\u53d6\u6271\u984d\uff08\u5343\u5186\uff09",
        }
    )
    st.dataframe(table, use_container_width=True, hide_index=True)
    st.caption(
        "\u91d1\u984d\u5358\u4f4d\u306f\u300c\u5343\u5186\u300d\u3067\u7d71\u4e00\u3057\u3066\u3044\u307e\u3059\u3002"
    )
    st.caption(
        "\u51fa\u5178\uff1a\u89b3\u5149\u5e81\u300e\u65c5\u884c\u696d\u8005\u53d6\u6271\u984d\u300f"
        "\uff08\u5404\u793e\u5225\u5185\u8a33Excel\u3092\u53d6\u5f97\u3057\u3066\u6574\u5f62\uff09"
    )
    st.caption(
        "\u30c7\u30fc\u30bf\u306f\u6bce\u9031\u81ea\u52d5\u66f4\u65b0\u3067\u3059\u3002"
        "\u53d6\u5f97\u5143\u30b5\u30a4\u30c8\u306e\u69cb\u9020\u5909\u66f4\u7b49\u306b\u3088\u308a"
        "\u66f4\u65b0\u304c\u9045\u308c\u308b\u5834\u5408\u304c\u3042\u308a\u307e\u3059\u3002"
    )


def load_airport_volume_meta() -> dict:
    if not AIRPORT_VOLUME_META_PATH.exists():
        return {}
    return json.loads(AIRPORT_VOLUME_META_PATH.read_text(encoding="utf-8"))


@st.cache_data(show_spinner=False)
def load_airport_volume_data() -> pd.DataFrame:
    if not SQLITE_PATH.exists():
        return pd.DataFrame()
    try:
        with sqlite3.connect(str(SQLITE_PATH)) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {AIRPORT_VOLUME_TABLE_NAME}", conn)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    for col in [
        "period_key",
        "airport_name_raw",
        "airport_name",
        "airport_code",
        "unit",
        "source_name",
        "source_url",
        "updated_at_utc",
    ]:
        if col in df.columns:
            df[col] = df[col].astype(str)
    if "arrivals" in df.columns:
        df["arrivals"] = pd.to_numeric(df["arrivals"], errors="coerce")

    return df


def month_to_quarter_period_key(period_key: str) -> str:
    m = re.fullmatch(r"(\d{4})-(\d{2})", str(period_key))
    if not m:
        return str(period_key)
    year = int(m.group(1))
    month = int(m.group(2))
    quarter = (month - 1) // 3 + 1
    return f"{year}Q{quarter}"


def render_airport_volume_view() -> None:
    st.title(
        "\u7a7a\u6e2f\u5225\u5165\u56fd\u8005\u6570\uff08\u30dc\u30ea\u30e5\u30fc\u30e0\uff09"
    )

    meta = load_airport_volume_meta()
    if meta:
        period_min = meta.get("period_min", "-")
        period_max = meta.get("period_max", "-")
        st.caption(
            f"\u6700\u7d42\u78ba\u8a8d\uff08UTC\uff09: {meta.get('fetched_at_utc')} / "
            f"\u5bfe\u8c61: {period_min} - {period_max}"
        )

    df = load_airport_volume_data()
    if df.empty:
        st.info(
            "\u7a7a\u6e2f\u5225\u5165\u56fd\u8005\u6570\u30c7\u30fc\u30bf\u306f\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
        return

    period_options = sorted(
        df["period_key"].dropna().astype(str).unique().tolist(),
        key=parse_period_sort_key,
    )
    if not period_options:
        st.warning(
            "\u671f\u9593\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    default_from_index = max(0, len(period_options) - 12)
    default_to_index = len(period_options) - 1

    col1, col2, col3 = st.columns([2, 2, 3])
    with col1:
        period_from = st.selectbox(
            "\u671f\u9593 From\uff08\u6708\u6b21\uff09",
            options=period_options,
            index=default_from_index,
            key="airport_volume_period_from",
        )
    with col2:
        period_to = st.selectbox(
            "\u671f\u9593 To\uff08\u6708\u6b21\uff09",
            options=period_options,
            index=default_to_index,
            key="airport_volume_period_to",
        )
    with col3:
        aggregate_mode = st.radio(
            "\u96c6\u8a08",
            options=["\u6708\u6b21", "\u56db\u534a\u671f"],
            horizontal=True,
            key="airport_volume_aggregate_mode",
        )

    if parse_period_sort_key(period_from) > parse_period_sort_key(period_to):
        period_from, period_to = period_to, period_from
        st.info(
            "From/To \u306e\u9806\u5e8f\u304c\u9006\u306e\u305f\u3081\u3001\u5165\u308c\u66ff\u3048\u3066\u8868\u793a\u3057\u3066\u3044\u307e\u3059\u3002"
        )

    range_df = df[
        (df["period_key"] >= period_from) & (df["period_key"] <= period_to)
    ].copy()
    if range_df.empty:
        st.info(
            "\u9078\u629e\u671f\u9593\u306b\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    airport_totals = (
        range_df.groupby(["airport_code", "airport_name"], as_index=False)["arrivals"]
        .sum()
        .sort_values("arrivals", ascending=False)
    )
    airport_options = airport_totals["airport_name"].astype(str).tolist()
    default_airports = airport_options[: min(6, len(airport_options))]
    selected_airports = st.multiselect(
        "\u7a7a\u6e2f",
        options=airport_options,
        default=default_airports,
        key="airport_volume_airports",
    )
    if not selected_airports:
        st.info(
            "\u7a7a\u6e2f\u3092\u9078\u629e\u3057\u3066\u304f\u3060\u3055\u3044\u3002"
        )
        return

    selected_df = range_df[range_df["airport_name"].isin(selected_airports)].copy()
    if selected_df.empty:
        st.info(
            "\u9078\u629e\u3057\u305f\u7a7a\u6e2f\u306e\u30c7\u30fc\u30bf\u304c\u3042\u308a\u307e\u305b\u3093\u3002"
        )
        return

    selected_df["period_key"] = selected_df["period_key"].astype(str)
    if aggregate_mode == "\u56db\u534a\u671f":
        selected_df["period_key"] = selected_df["period_key"].map(
            month_to_quarter_period_key
        )

    chart_df = (
        selected_df.groupby(["period_key", "airport_name"], as_index=False)["arrivals"]
        .sum()
        .copy()
    )
    ranking_df = (
        selected_df.groupby(["airport_name"], as_index=False)["arrivals"]
        .sum()
        .sort_values("arrivals", ascending=False)
    )

    months_count = range_df["period_key"].nunique()
    total_arrivals = float(selected_df["arrivals"].sum())
    monthly_avg = total_arrivals / months_count if months_count > 0 else None

    kpi_col1, kpi_col2 = st.columns(2)
    kpi_col1.metric(
        "\u671f\u9593\u5408\u8a08\u5165\u56fd\u8005\u6570\uff08\u4eba\uff09",
        format_metric_value(total_arrivals),
    )
    kpi_col2.metric(
        "\u671f\u9593\u5e73\u5747\uff08\u6708\u6b21\uff09",
        format_metric_value(monthly_avg, 1),
    )

    st.subheader(
        "\u63a8\u79fb\uff08"
        + (
            "\u56db\u534a\u671f"
            if aggregate_mode == "\u56db\u534a\u671f"
            else "\u6708\u6b21"
        )
        + "\uff09"
    )
    top_n = st.slider(
        "TopN\uff08\u30c1\u30e3\u30fc\u30c8\u8868\u793a\u7a7a\u6e2f\u6570\uff09",
        min_value=3,
        max_value=12,
        value=min(6, len(selected_airports)),
        key="airport_volume_top_n",
    )
    chart_airports = set(ranking_df.head(top_n)["airport_name"].astype(str).tolist())
    chart_view = chart_df[chart_df["airport_name"].isin(chart_airports)].copy()

    if len(selected_airports) > top_n:
        st.caption(
            f"\u8868\u793a\u8ca0\u8377\u6291\u5236\u306e\u305f\u3081\u3001\u63a8\u79fb\u30c1\u30e3\u30fc\u30c8\u306f\u4e0a\u4f4d{top_n}\u7a7a\u6e2f\u306b\u9650\u5b9a\u3057\u3066\u3044\u307e\u3059\u3002"
        )

    if chart_view.empty:
        st.info(
            "\u63a8\u79fb\u30c1\u30e3\u30fc\u30c8\u7528\u306e\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        period_order = sorted(
            chart_view["period_key"].astype(str).unique().tolist(),
            key=parse_period_sort_key,
        )
        trend_chart = (
            alt.Chart(chart_view)
            .mark_line(point=True)
            .encode(
                x=alt.X(
                    "period_key:N",
                    sort=period_order,
                    title="\u671f\u9593\uff08\u56db\u534a\u671f\uff09"
                    if aggregate_mode == "\u56db\u534a\u671f"
                    else "\u671f\u9593\uff08\u6708\u6b21\uff09",
                ),
                y=alt.Y(
                    "arrivals:Q", title="\u5165\u56fd\u8005\u6570\uff08\u4eba\uff09"
                ),
                color=alt.Color("airport_name:N", title="\u7a7a\u6e2f"),
                tooltip=[
                    alt.Tooltip("period_key:N", title="\u671f\u9593"),
                    alt.Tooltip("airport_name:N", title="\u7a7a\u6e2f"),
                    alt.Tooltip(
                        "arrivals:Q", title="\u5165\u56fd\u8005\u6570", format=",.0f"
                    ),
                ],
            )
            .properties(height=360)
        )
        st.altair_chart(trend_chart, use_container_width=True)

    st.subheader("\u7a7a\u6e2f\u5225\u30e9\u30f3\u30ad\u30f3\u30b0")
    ranking_top = ranking_df.head(top_n).copy()
    if ranking_top.empty:
        st.info(
            "\u30e9\u30f3\u30ad\u30f3\u30b0\u30c7\u30fc\u30bf\u304c\u672a\u53d6\u5f97\u3067\u3059\u3002"
        )
    else:
        ranking_chart = (
            alt.Chart(ranking_top)
            .mark_bar()
            .encode(
                x=alt.X(
                    "arrivals:Q",
                    title="\u671f\u9593\u5408\u8a08\u5165\u56fd\u8005\u6570\uff08\u4eba\uff09",
                ),
                y=alt.Y("airport_name:N", sort="-x", title="\u7a7a\u6e2f"),
                tooltip=[
                    alt.Tooltip("airport_name:N", title="\u7a7a\u6e2f"),
                    alt.Tooltip(
                        "arrivals:Q", title="\u5165\u56fd\u8005\u6570", format=",.0f"
                    ),
                ],
            )
            .properties(height=360)
        )
        st.altair_chart(ranking_chart, use_container_width=True)

    st.subheader("\u8868")
    table_df = chart_df.copy()
    table_df["_period_sort"] = table_df["period_key"].map(parse_period_sort_key)
    table_df = (
        table_df.sort_values(["_period_sort", "airport_name"], ascending=[True, True])
        .drop(columns=["_period_sort"])
        .reset_index(drop=True)
    )
    st.dataframe(
        table_df[["period_key", "airport_name", "arrivals"]],
        use_container_width=True,
        hide_index=True,
    )

    csv_data = table_df[["period_key", "airport_name", "arrivals"]].to_csv(index=False)
    st.download_button(
        "\u8868\u3092CSV\u30c0\u30a6\u30f3\u30ed\u30fc\u30c9",
        data=csv_data.encode("utf-8-sig"),
        file_name=f"airport_arrivals_{period_from}_{period_to}.csv",
        mime="text/csv",
        key="airport_volume_csv_download",
    )
    st.caption(
        "\u51fa\u5178\uff1ae-Stat\u300e\u51fa\u5165\u56fd\u7ba1\u7406\u7d71\u8a08\u300f\u7dcf\u62ec\uff08\u6e2f\u5225 \u51fa\u5165\u56fd\u8005\uff09"
        "\uff08\u6708\u6b21Excel\u3092\u53d6\u5f97\u3057\u3066\u6574\u5f62\uff09"
    )
    st.caption(
        "\u30c7\u30fc\u30bf\u306f\u6bce\u9031\u81ea\u52d5\u66f4\u65b0\u3067\u3059\u3002"
        "\u53d6\u5f97\u5143\u30b5\u30a4\u30c8\u306e\u69cb\u9020\u5909\u66f4\u7b49\u306b\u3088\u308a"
        "\u66f4\u65b0\u304c\u9045\u308c\u308b\u5834\u5408\u304c\u3042\u308a\u307e\u3059\u3002"
    )


# ---------------------------------------------------------------------------
# Events Hub
# ---------------------------------------------------------------------------
EVENT_CATEGORY_ALL = "すべて"
EVENT_CATEGORY_CONCERT = "コンサート（その他含む）"
EVENT_CATEGORY_BASEBALL = "野球"
EVENT_CATEGORY_OPTIONS = [
    EVENT_CATEGORY_ALL,
    EVENT_CATEGORY_CONCERT,
    EVENT_CATEGORY_BASEBALL,
]


def classify_event_category(title: str, performers: str, description: str) -> str:
    text = " ".join([str(title or ""), str(performers or ""), str(description or "")])
    normalized_text = text.lower()

    baseball_keywords = [
        "野球",
        "ベースボール",
        "baseball",
        "npb",
        "プロ野球",
        "甲子園",
        "侍ジャパン",
        "オープン戦",
        "公式戦",
        "ファーム",
        "クライマックスシリーズ",
        "日本シリーズ",
        "セ・リーグ",
        "パ・リーグ",
        "交流戦",
    ]
    baseball_team_keywords = [
        "オリックス",
        "バファローズ",
        "阪神",
        "タイガース",
        "巨人",
        "ジャイアンツ",
        "ヤクルト",
        "スワローズ",
        "広島",
        "カープ",
        "中日",
        "ドラゴンズ",
        "dena",
        "ベイスターズ",
        "ベイス",
        "ソフトバンク",
        "ホークス",
        "日本ハム",
        "日ハム",
        "ファイターズ",
        "楽天",
        "イーグルス",
        "西武",
        "ライオンズ",
        "ロッテ",
        "マリーンズ",
    ]
    versus_keywords = ["vs", "ｖｓ", "対", " 対戦 "]
    if any(keyword in normalized_text for keyword in baseball_keywords):
        return EVENT_CATEGORY_BASEBALL
    has_team_keyword = any(
        keyword in normalized_text for keyword in baseball_team_keywords
    )
    has_versus = any(keyword in normalized_text for keyword in versus_keywords)
    if has_team_keyword and has_versus:
        return EVENT_CATEGORY_BASEBALL

    concert_keywords = [
        "ライブ",
        "コンサート",
        "公演",
        "ツアー",
        "フェス",
        "リサイタル",
        "オーケストラ",
        "music",
        "concert",
        "festival",
        "band",
        "gig",
        "live",
        "tour",
        "dome",
        "arena",
        "hall",
        "zepp",
        "oneman",
        "one man",
        "world tour",
        "fan meeting",
        "showcase",
        "弾き語り",
        "ワンマン",
        "ワールドツアー",
        "ドームツアー",
        "アリーナツアー",
    ]
    if any(keyword in normalized_text for keyword in concert_keywords):
        return EVENT_CATEGORY_CONCERT
    return EVENT_CATEGORY_CONCERT


@st.cache_data(show_spinner=False)
def load_events_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load events + venues from events.sqlite. Returns (df_events, df_venues)."""
    if not EVENTS_DB_PATH.exists():
        return pd.DataFrame(), pd.DataFrame()
    try:
        with sqlite3.connect(str(EVENTS_DB_PATH)) as conn:
            df_events = pd.read_sql_query("SELECT * FROM events", conn)
            df_venues = pd.read_sql_query("SELECT * FROM venues", conn)
        return df_events, df_venues
    except Exception:
        return pd.DataFrame(), pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_event_signals_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load signals + signal_sources from event_signals.sqlite."""
    if not EVENT_SIGNALS_DB_PATH.exists():
        return pd.DataFrame(), pd.DataFrame()
    try:
        with sqlite3.connect(str(EVENT_SIGNALS_DB_PATH)) as conn:
            df_signals = pd.read_sql_query("SELECT * FROM signals", conn)
            df_sources = pd.read_sql_query("SELECT * FROM signal_sources", conn)
        return df_signals, df_sources
    except Exception:
        return pd.DataFrame(), pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_events_artist_inferred_map(
) -> dict[str, tuple[str, str]]:
    if not EVENTS_ARTIST_INFERRED_PATH.exists():
        return {}
    try:
        df = pd.read_csv(EVENTS_ARTIST_INFERRED_PATH, dtype=str).fillna("")
    except Exception:
        return {}
    out: dict[str, tuple[str, str]] = {}
    for _, row in df.iterrows():
        title = str(row.get("title", "")).strip()
        artist_name = str(row.get("artist_name", "")).strip()
        confidence = str(row.get("artist_confidence", "")).strip().lower() or "high"
        if not title or not artist_name:
            continue
        out[title] = (artist_name, confidence)
    return out


def render_event_signals_view() -> None:
    st.title("全国イベント速報（ニュース）")

    df_signals, df_sources = load_event_signals_data()
    if df_signals.empty:
        st.error(
            "イベント速報データがありません。"
            "`uv run python -m scripts.update_event_signals_data` を実行してください。"
        )
        return

    source_master = df_sources[["source_id", "source_name"]].drop_duplicates()
    df = df_signals.merge(source_master, on="source_id", how="left")

    published_utc = pd.to_datetime(df["published_at_utc"], utc=True, errors="coerce")
    df = df[published_utc.notna()].copy()
    if df.empty:
        st.warning("表示可能な速報データがありません。")
        return

    df["published_dt_utc"] = published_utc[published_utc.notna()]
    df["published_dt_jst"] = df["published_dt_utc"].dt.tz_convert("Asia/Tokyo")
    df["published_jst"] = df["published_dt_jst"].dt.strftime("%Y-%m-%d %H:%M")
    df["score"] = pd.to_numeric(df.get("score"), errors="coerce").fillna(0).astype(int)

    def _load_labels(labels_json: object) -> dict[str, object]:
        if not isinstance(labels_json, str) or not labels_json.strip():
            return {}
        try:
            parsed = json.loads(labels_json)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}

    labels_series = df.get(
        "labels_json", pd.Series(index=df.index, dtype=object)
    ).apply(_load_labels)

    df["raw_artist_name"] = labels_series.apply(
        lambda d: str(d.get("artist_name", "")).strip() if isinstance(d, dict) else ""
    )
    df["raw_venue_name"] = labels_series.apply(
        lambda d: str(d.get("venue_name", "")).strip() if isinstance(d, dict) else ""
    )
    df["raw_pref_name"] = labels_series.apply(
        lambda d: (
            str(d.get("pref_name", d.get("raw_pref_name", ""))).strip()
            if isinstance(d, dict)
            else ""
        )
    )
    df["event_info"] = labels_series.apply(
        lambda d: str(d.get("event_info", "")).strip() if isinstance(d, dict) else ""
    )
    df["artist_confidence"] = labels_series.apply(
        lambda d: (
            str(d.get("artist_confidence", "")).strip().lower()
            if isinstance(d, dict)
            else ""
        )
    )
    df["artist_confidence"] = df["artist_confidence"].replace("", "low")

    def _parse_iso_date(value: object):
        if not isinstance(value, str) or not value.strip():
            return pd.NaT
        return pd.to_datetime(value, format="%Y-%m-%d", errors="coerce")

    event_start_parsed = labels_series.apply(
        lambda d: _parse_iso_date(d.get("event_start_date"))
    )
    event_end_parsed = labels_series.apply(
        lambda d: _parse_iso_date(d.get("event_end_date"))
    )
    df["raw_event_start_dt"] = event_start_parsed
    published_date_naive = df["published_dt_jst"].dt.tz_localize(None).dt.normalize()

    df["event_start_dt"] = event_start_parsed.where(
        event_start_parsed.notna(), published_date_naive
    )
    df["event_end_dt"] = event_end_parsed.where(
        event_end_parsed.notna(), df["event_start_dt"]
    )
    df["event_start_date"] = df["event_start_dt"].dt.date
    df["event_end_date"] = df["event_end_dt"].dt.date
    df["event_start_ym"] = df["event_start_dt"].dt.strftime("%Y-%m")
    df["event_end_ym"] = df["event_end_dt"].dt.strftime("%Y-%m")
    df["event_date_jst"] = df.apply(
        lambda row: (
            row["event_start_date"].strftime("%Y-%m-%d")
            if row["event_start_date"] == row["event_end_date"]
            else f"{row['event_start_date'].strftime('%Y-%m-%d')} ～ {row['event_end_date'].strftime('%Y-%m-%d')}"
        ),
        axis=1,
    )

    df["artist_name"] = df["raw_artist_name"]
    df["venue_name"] = df["raw_venue_name"]

    df["artist_name"] = df.apply(
        lambda row: row["artist_name"] if row["artist_name"] else row["source_name"],
        axis=1,
    )
    df["venue_name"] = df.apply(
        lambda row: row["venue_name"] if row["venue_name"] else "-",
        axis=1,
    )
    df["event_info"] = df.apply(
        lambda row: row["event_info"] if row["event_info"] else row["snippet"],
        axis=1,
    )
    df["published_label"] = df["published_jst"]

    def _normalize_hhmm(value: object) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        match = re.search(r"(?<!\d)(\d{1,2}:\d{2})(?!\d)", text)
        if not match:
            return ""
        hour = int(match.group(1).split(":")[0])
        minute = int(match.group(1).split(":")[1])
        if not (0 <= hour <= 23 and 0 <= minute <= 59):
            return ""
        return f"{hour:02d}:{minute:02d}"

    def _extract_start_time_from_event_info(value: object) -> str:
        if not isinstance(value, str):
            return ""
        text = value.strip()
        if not text:
            return ""
        for pattern in [
            r"(?<!\d)(\d{1,2}:\d{2})(?!\d)\s*開演",
            r"\d{4}-\d{2}-\d{2}\s+(\d{1,2}:\d{2})",
            r"(?<!\d)(\d{1,2}:\d{2})(?!\d)",
        ]:
            match = re.search(pattern, text)
            if match:
                return _normalize_hhmm(match.group(1))
        return ""

    df["raw_start_time_jst"] = labels_series.apply(
        lambda d: (
            _normalize_hhmm(
                d.get("event_start_time", d.get("start_time", d.get("event_time", "")))
            )
            if isinstance(d, dict)
            else ""
        )
    )
    df["start_time_jst"] = df["raw_start_time_jst"]
    missing_start_time_mask = df["start_time_jst"].eq("")
    if missing_start_time_mask.any():
        df.loc[missing_start_time_mask, "start_time_jst"] = df.loc[
            missing_start_time_mask, "event_info"
        ].apply(_extract_start_time_from_event_info)
    df["start_time_jst"] = df["start_time_jst"].replace("", "-")

    def _normalize_pref_name(value: object) -> str:
        pref = str(value).strip() if value is not None else ""
        if not pref:
            return ""
        if pref in ("東京", "東京都"):
            return "東京都"
        if pref in ("大阪", "大阪府"):
            return "大阪府"
        if pref in ("京都", "京都府"):
            return "京都府"
        if pref == "北海道":
            return "北海道"
        if pref.endswith(("都", "道", "府", "県")):
            return pref
        return f"{pref}県"

    def _extract_pref_from_event_info(value: object) -> str:
        if not isinstance(value, str):
            return ""
        match = re.match(r"^\s*[\[［]\s*([^\]］]+?)\s*[\]］]", value.strip())
        if not match:
            return ""
        return match.group(1).strip()

    df["pref_name"] = df["raw_pref_name"].apply(_normalize_pref_name)
    missing_pref_mask = df["pref_name"].eq("")
    if missing_pref_mask.any():
        df.loc[missing_pref_mask, "pref_name"] = (
            df.loc[missing_pref_mask, "event_info"]
            .apply(_extract_pref_from_event_info)
            .apply(_normalize_pref_name)
        )

    ym_options = sorted(
        {
            *df["event_start_ym"].dropna().astype(str).tolist(),
            *df["event_end_ym"].dropna().astype(str).tolist(),
        }
    )
    if not ym_options:
        st.warning("表示可能な速報データがありません。")
        return

    min_ym = ym_options[0]
    max_ym = ym_options[-1]
    today_jst = pd.Timestamp.now(tz="Asia/Tokyo").date()
    default_start_ym = build_ym(today_jst.year, today_jst.month)
    default_end_ym = max_ym
    default_start_ym, _ = clamp_ym_to_available_range(default_start_ym, min_ym, max_ym)
    default_end_ym, _ = clamp_ym_to_available_range(default_end_ym, min_ym, max_ym)

    year_options = sorted({int(ym[:4]) for ym in ym_options})
    month_options = list(range(1, 13))

    def _fmt_month(month: int) -> str:
        return f"{month:02d}"

    default_start_year = int(default_start_ym[:4])
    default_start_month = int(default_start_ym[5:7])
    default_end_year = int(default_end_ym[:4])
    default_end_month = int(default_end_ym[5:7])

    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1, 2, 1])
    with col_f1:
        start_year = st.selectbox(
            "開始（年）",
            options=year_options,
            index=year_options.index(default_start_year),
            key="signals_start_year",
        )
    with col_f2:
        start_month = st.selectbox(
            "開始（月）",
            options=month_options,
            index=month_options.index(default_start_month),
            format_func=_fmt_month,
            key="signals_start_month",
        )
    with col_f3:
        end_year = st.selectbox(
            "終了（年）",
            options=year_options,
            index=year_options.index(default_end_year),
            key="signals_end_year",
        )
    with col_f4:
        end_month = st.selectbox(
            "終了（月）",
            options=month_options,
            index=month_options.index(default_end_month),
            format_func=_fmt_month,
            key="signals_end_month",
        )

    ym_from = build_ym(int(start_year), int(start_month))
    ym_to = build_ym(int(end_year), int(end_month))
    ym_from, from_clamped = clamp_ym_to_available_range(ym_from, min_ym, max_ym)
    ym_to, to_clamped = clamp_ym_to_available_range(ym_to, min_ym, max_ym)
    if from_clamped:
        st.warning(f"開始年月をデータ範囲に合わせて {ym_from} に補正しました。")
    if to_clamped:
        st.warning(f"終了年月をデータ範囲に合わせて {ym_to} に補正しました。")
    if ym_to_int(ym_from) > ym_to_int(ym_to):
        ym_from, ym_to = ym_to, ym_from
        st.warning(
            f"イベント年月が逆順だったため、{ym_from} ～ {ym_to} に入れ替えました。"
        )

    source_options = (
        df[["source_id", "source_name"]]
        .drop_duplicates()
        .sort_values("source_name")
        .to_dict("records")
    )
    source_name_to_id = {
        str(row["source_name"]): str(row["source_id"]) for row in source_options
    }
    selected_source_names = st.multiselect(
        "ソース",
        options=list(source_name_to_id.keys()),
        default=list(source_name_to_id.keys()),
        key="signals_sources",
    )
    selected_source_ids = [source_name_to_id[name] for name in selected_source_names]

    pref_options = sorted(
        [
            pref
            for pref in df["pref_name"].dropna().astype(str).unique().tolist()
            if pref.strip()
        ]
    )
    selected_prefs = render_pref_toggles_two_step(
        pref_options=pref_options,
        pref_selection_key="signals_pref",
        pref_toggle_key_prefix="signals_pref_toggle",
        region_selection_key="signals_regions",
        region_toggle_key_prefix="signals_region_toggle",
    )

    keyword = st.text_input("キーワード（タイトル/抜粋）", key="signals_keyword")
    sort_label = st.radio(
        "並び順",
        ["掲載日時（新しい順）", "イベント日（早い順）"],
        horizontal=True,
        key="signals_sort",
    )
    mask = (df["event_end_ym"] >= ym_from) & (df["event_start_ym"] <= ym_to)
    if selected_source_ids:
        mask &= df["source_id"].isin(selected_source_ids)
    if selected_prefs:
        mask &= df["pref_name"].isin(selected_prefs)
    if keyword:
        keyword_lower = keyword.lower()
        mask &= df["title"].fillna("").str.lower().str.contains(keyword_lower) | df[
            "event_info"
        ].fillna("").str.lower().str.contains(keyword_lower)

    filtered = df[mask].copy()
    if sort_label == "掲載日時（新しい順）":
        filtered = filtered.sort_values(
            ["published_dt_jst", "event_start_dt", "title"],
            ascending=[False, True, True],
        )
    else:
        filtered = filtered.sort_values(
            ["event_start_dt", "event_end_dt", "title"], ascending=[True, True, True]
        )
    now_jst = pd.Timestamp.now(tz="Asia/Tokyo")
    filtered["is_new_24h"] = filtered["published_dt_jst"] >= (
        now_jst - pd.Timedelta(hours=24)
    )
    filtered["new_badge"] = filtered["is_new_24h"].map(lambda v: "NEW" if bool(v) else "")

    new_count = int(filtered["is_new_24h"].sum()) if not filtered.empty else 0
    st.markdown(f"**{len(filtered)}** 件の速報")
    if new_count > 0:
        st.success(f"新規イベント（直近24時間）: {new_count} 件")
    else:
        st.caption("新規イベント（直近24時間）: 0 件")

    table_df = filtered[
        [
            "new_badge",
            "event_date_jst",
            "start_time_jst",
            "venue_name",
            "pref_name",
            "artist_name",
            "title",
            "published_label",
            "source_name",
            "url",
            "event_info",
        ]
    ].rename(
        columns={
            "new_badge": "新着",
            "event_date_jst": "イベント日",
            "start_time_jst": "開始時間",
            "artist_name": "アーティスト",
            "venue_name": "会場",
            "pref_name": "都道府県",
            "published_label": "掲載日",
            "source_name": "ソース",
            "title": "タイトル",
            "url": "URL",
            "event_info": "イベント情報",
        }
    )
    st.dataframe(
        table_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "URL": st.column_config.LinkColumn("URL", display_text="リンク"),
        },
    )

    st.subheader("エクスポート")
    export_cols = [
        "signal_uid",
        "source_id",
        "source_name",
        "artist_name",
        "venue_name",
        "pref_name",
        "event_date_jst",
        "start_time_jst",
        "published_label",
        "is_new_24h",
        "published_at_utc",
        "score",
        "title",
        "url",
        "event_info",
        "snippet",
        "labels_json",
        "updated_at_utc",
    ]
    export_df = filtered[export_cols].copy()

    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        csv_data = export_df.to_csv(index=False)
        st.download_button(
            "CSVダウンロード",
            data=csv_data.encode("utf-8-sig"),
            file_name=f"event_signals_{ym_from}_{ym_to}.csv",
            mime="text/csv",
            key="signals_csv_dl",
            use_container_width=True,
        )
    with dl_col2:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="event_signals")
        st.download_button(
            "Excelダウンロード",
            data=buf.getvalue(),
            file_name=f"event_signals_{ym_from}_{ym_to}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="signals_excel_dl",
            use_container_width=True,
        )

    st.caption(
        "速報データは本文を保存せず、掲載日時・タイトル・URL・短い抜粋のみ保持します。"
    )
    st.caption(
        "データは1日2回（12時間ごと）自動更新です。"
        "取得元サイトの構造変更等により更新が遅れる場合があります。"
    )


def render_events_view() -> None:
    st.title("全国イベント情報（会場公式）")

    df_events, df_venues = load_events_data()
    if df_events.empty:
        st.error(
            "イベントデータがありません。"
            "`uv run python -m scripts.update_events_data` を実行してください。"
        )
        return

    # Merge venue info
    df = df_events.merge(
        df_venues[
            ["venue_id", "venue_name", "pref_code", "pref_name", "capacity"]
        ].rename(columns={"capacity": "venue_capacity"}),
        on="venue_id",
        how="left",
    )
    df["display_capacity"] = df["capacity"].fillna(df["venue_capacity"])
    df["pref_code"] = df["pref_code"].astype(str).str.zfill(2)
    df["event_category"] = df.apply(
        lambda row: classify_event_category(
            row.get("title"), row.get("performers"), row.get("description")
        ),
        axis=1,
    )
    # --- Filters ---
    from datetime import date as dt_date
    from datetime import timedelta

    df["event_ym"] = df["start_date"].astype(str).str.slice(0, 7)
    ym_options = sorted(
        [
            ym
            for ym in df["event_ym"].dropna().astype(str).unique().tolist()
            if re.fullmatch(r"\d{4}-\d{2}", ym)
        ]
    )
    if not ym_options:
        st.info("期間選択に使える年月データがありません。")
        return

    min_ym = ym_options[0]
    max_ym = ym_options[-1]
    today = dt_date.today()
    default_from_ym = build_ym(today.year, today.month)
    default_to_date = today + timedelta(days=180)
    default_to_ym = build_ym(default_to_date.year, default_to_date.month)
    default_from_ym, _ = clamp_ym_to_available_range(default_from_ym, min_ym, max_ym)
    default_to_ym, _ = clamp_ym_to_available_range(default_to_ym, min_ym, max_ym)

    year_options = sorted({int(ym[:4]) for ym in ym_options})
    month_options = list(range(1, 13))

    def _fmt_month(month: int) -> str:
        return f"{month:02d}"

    default_from_year = int(default_from_ym[:4])
    default_from_month = int(default_from_ym[5:7])
    default_to_year = int(default_to_ym[:4])
    default_to_month = int(default_to_ym[5:7])

    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1, 2, 1])
    with col_f1:
        start_year = st.selectbox(
            "開始（年）",
            options=year_options,
            index=year_options.index(default_from_year),
            key="events_start_year",
        )
    with col_f2:
        start_month = st.selectbox(
            "開始（月）",
            options=month_options,
            index=month_options.index(default_from_month),
            format_func=_fmt_month,
            key="events_start_month",
        )
    with col_f3:
        end_year = st.selectbox(
            "終了（年）",
            options=year_options,
            index=year_options.index(default_to_year),
            key="events_end_year",
        )
    with col_f4:
        end_month = st.selectbox(
            "終了（月）",
            options=month_options,
            index=month_options.index(default_to_month),
            format_func=_fmt_month,
            key="events_end_month",
        )

    ym_from = build_ym(int(start_year), int(start_month))
    ym_to = build_ym(int(end_year), int(end_month))
    ym_from, from_clamped = clamp_ym_to_available_range(ym_from, min_ym, max_ym)
    ym_to, to_clamped = clamp_ym_to_available_range(ym_to, min_ym, max_ym)
    if from_clamped:
        st.warning(f"開始年月をデータ範囲に合わせて {ym_from} に補正しました。")
    if to_clamped:
        st.warning(f"終了年月をデータ範囲に合わせて {ym_to} に補正しました。")
    if ym_to_int(ym_from) > ym_to_int(ym_to):
        ym_from, ym_to = ym_to, ym_from
        st.warning(
            f"開始年月と終了年月が逆だったため、{ym_from} ～ {ym_to} に入れ替えました。"
        )

    # Pref filter (横並びトグル・複数選択)
    pref_master = (
        df[["pref_code", "pref_name"]]
        .dropna(subset=["pref_code", "pref_name"])
        .drop_duplicates()
        .sort_values("pref_code")
    )
    pref_options = pref_master["pref_name"].astype(str).tolist()
    selected_prefs = render_pref_toggles_two_step(
        pref_options=pref_options,
        pref_selection_key="events_pref",
        pref_toggle_key_prefix="events_pref_toggle",
        region_selection_key="events_regions",
        region_toggle_key_prefix="events_region_toggle",
    )

    # Venue filter (narrowed by pref)
    venue_pool = df.copy()
    if selected_prefs:
        venue_pool = venue_pool[venue_pool["pref_name"].isin(selected_prefs)]
    venue_options = sorted(venue_pool["venue_name"].dropna().unique().tolist())
    current_selected_venues = [
        str(v)
        for v in st.session_state.get("events_venue", [])
        if str(v) in venue_options
    ]
    if current_selected_venues != st.session_state.get("events_venue", []):
        st.session_state["events_venue"] = current_selected_venues
    selected_venues = st.multiselect("会場", venue_options, key="events_venue")

    # Keyword
    keyword = st.text_input("キーワード（タイトル/出演者/説明）", key="events_keyword")
    selected_category = st.radio(
        "種別", EVENT_CATEGORY_OPTIONS, horizontal=True, key="events_category"
    )

    # Status
    include_cancelled = st.checkbox(
        "cancelled/postponed を含む", value=False, key="events_incl_cancel"
    )

    # --- Apply filters ---
    mask = (df["event_ym"] >= ym_from) & (df["event_ym"] <= ym_to)
    if selected_prefs:
        mask &= df["pref_name"].isin(selected_prefs)
    if selected_venues:
        mask &= df["venue_name"].isin(selected_venues)
    if keyword:
        kw_lower = keyword.lower()
        kw_mask = (
            df["title"].str.lower().str.contains(kw_lower, na=False)
            | df["performers"].fillna("").str.lower().str.contains(kw_lower, na=False)
            | df["description"].fillna("").str.lower().str.contains(kw_lower, na=False)
        )
        mask &= kw_mask
    if selected_category != EVENT_CATEGORY_ALL:
        mask &= df["event_category"] == selected_category
    if not include_cancelled:
        mask &= df["status"].isin(["scheduled", "unknown"])
    filtered = df[mask].copy().sort_values("start_date").reset_index(drop=True)
    filtered["artist_name"] = filtered["performers"].fillna("").astype(str).str.strip()
    filtered["artist_confidence"] = filtered["artist_name"].map(
        lambda v: "source" if v else "low"
    )
    missing_artist_mask = filtered["artist_name"].eq("")
    if missing_artist_mask.any():
        inferred_map = load_events_artist_inferred_map()
        inferred = filtered.loc[missing_artist_mask, "title"].map(
            lambda title: inferred_map.get(str(title).strip(), ("", "low"))
        )
        filtered.loc[missing_artist_mask, "artist_name"] = inferred.apply(
            lambda row: row[0]
        )
        filtered.loc[missing_artist_mask, "artist_confidence"] = inferred.apply(
            lambda row: row[1]
        )

    st.markdown(f"**{len(filtered)}** 件のイベント")

    # --- Table ---
    display_cols = [
        "start_date",
        "start_time",
        "venue_name",
        "pref_name",
        "title",
        "artist_name",
        "status",
        "display_capacity",
        "url",
    ]
    display_rename = {
        "start_date": "イベント日",
        "start_time": "開始時間",
        "venue_name": "会場",
        "pref_name": "都道府県",
        "title": "タイトル",
        "artist_name": "アーティスト",
        "status": "ステータス",
        "display_capacity": "キャパシティ",
        "url": "URL",
    }
    table_df = filtered[display_cols].rename(columns=display_rename)
    st.dataframe(
        table_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "URL": st.column_config.LinkColumn("URL", display_text="リンク"),
        },
    )

    # --- Chart: Event intensity ---
    st.subheader("イベント強度（日別）")
    if not filtered.empty:
        chart_df = filtered.copy()
        chart_df["display_capacity"] = (
            pd.to_numeric(chart_df["display_capacity"], errors="coerce")
            .fillna(0)
            .astype(int)
        )
        daily = (
            chart_df.groupby("start_date")
            .agg(
                event_count=("event_uid", "count"),
                total_capacity=("display_capacity", "sum"),
            )
            .reset_index()
        )
        daily.columns = ["日付", "イベント件数", "合計キャパシティ"]

        base = alt.Chart(daily).encode(
            x=alt.X("日付:T", title="日付"),
        )
        bars = base.mark_bar(opacity=0.6).encode(
            y=alt.Y("合計キャパシティ:Q", title="合計キャパシティ"),
            tooltip=["日付:T", "イベント件数:Q", "合計キャパシティ:Q"],
        )
        line = base.mark_line(color="red", strokeWidth=2).encode(
            y=alt.Y("イベント件数:Q", title="イベント件数"),
        )
        chart = (
            alt.layer(bars, line).resolve_scale(y="independent").properties(height=400)
        )
        st.altair_chart(cast(alt.Chart, chart), use_container_width=True)
    else:
        st.info("表示できるイベントがありません。")

    # --- Export ---
    st.subheader("エクスポート")
    export_cols = [
        "event_uid",
        "venue_id",
        "venue_name",
        "pref_code",
        "pref_name",
        "start_date",
        "start_time",
        "end_date",
        "end_time",
        "all_day",
        "title",
        "artist_name",
        "artist_confidence",
        "performers",
        "status",
        "display_capacity",
        "url",
        "source_type",
        "source_url",
        "updated_at_utc",
    ]
    export_rename = {"display_capacity": "capacity"}
    export_df = filtered[[c for c in export_cols if c in filtered.columns]].rename(
        columns=export_rename
    )

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        csv_data = export_df.to_csv(index=False)
        st.download_button(
            "CSVダウンロード",
            data=csv_data.encode("utf-8-sig"),
            file_name=f"events_{ym_from}_{ym_to}.csv",
            mime="text/csv",
            key="events_csv_dl",
            use_container_width=True,
        )
    with col_dl2:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="events")
        st.download_button(
            "Excelダウンロード",
            data=buf.getvalue(),
            file_name=f"events_{ym_from}_{ym_to}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="events_excel_dl",
            use_container_width=True,
        )

    st.caption(
        "データは毎週自動更新です。"
        "取得元サイトの構造変更等により更新が遅れる場合があります。"
    )


def main() -> None:
    st.set_page_config(
        page_title="\u5e02\u5834\u7d71\u8a08\u30d3\u30e5\u30fc\u30a2",
        page_icon="assets/logo_header.svg",
        layout="wide",
    )

    lp_url = "https://deltahelmlab.com/?utm_source=market_stats_viewer&utm_medium=app&utm_campaign=cross_link"
    with st.sidebar:
        st.markdown("### \u904b\u55b6\u5143")
        st.link_button(
            "DeltaHelm Lab\uff08\u516c\u5f0f\u30b5\u30a4\u30c8\uff09",
            lp_url,
            use_container_width=True,
        )
        st.caption(
            "\u30b5\u30fc\u30d3\u30b9\u8a73\u7d30\u30fb\u304a\u554f\u3044\u5408\u308f\u305b\u306f\u3053\u3061\u3089"
        )
        # --- 統計データ / 参考情報 の排他 radio ---
        _STATS_OPTIONS = [
            DATASET_LABEL_STAY,
            DATASET_LABEL_TCD,
            DATASET_LABEL_ICD,
            DATASET_LABEL_AIRPORT_VOLUME,
            DATASET_LABEL_TA,
        ]
        _REF_OPTIONS = [
            DATASET_LABEL_EVENTS_OFFICIAL,
            DATASET_LABEL_EVENTS_SIGNALS,
        ]

        if "_active_dataset" not in st.session_state:
            st.session_state["_active_dataset"] = _STATS_OPTIONS[0]

        active = st.session_state["_active_dataset"]
        _is_ref = active in _REF_OPTIONS

        def _on_stats_change():
            st.session_state["_active_dataset"] = st.session_state["_stats_sel"]
            st.session_state.pop("_ref_sel", None)

        def _on_ref_change():
            st.session_state["_active_dataset"] = st.session_state["_ref_sel"]
            st.session_state.pop("_stats_sel", None)

        st.radio(
            "統計データ",
            _STATS_OPTIONS,
            index=None if _is_ref else _STATS_OPTIONS.index(active),
            key="_stats_sel",
            on_change=_on_stats_change,
        )
        st.markdown("---")
        st.radio(
            "参考情報",
            _REF_OPTIONS,
            index=_REF_OPTIONS.index(active) if _is_ref else None,
            key="_ref_sel",
            on_change=_on_ref_change,
        )

        dataset_type = st.session_state["_active_dataset"]

    if dataset_type == DATASET_LABEL_STAY:
        render_stay_stats_view()
        return

    if dataset_type == DATASET_LABEL_TCD:
        render_tcd_view()
        return

    if dataset_type == DATASET_LABEL_ICD:
        render_icd_view()
        return

    if dataset_type == DATASET_LABEL_TA:
        render_ta_view()
        return

    if dataset_type == DATASET_LABEL_EVENTS_OFFICIAL:
        render_events_view()
        return

    if dataset_type == DATASET_LABEL_EVENTS_SIGNALS:
        render_event_signals_view()
        return

    render_airport_volume_view()


if __name__ == "__main__":
    main()
